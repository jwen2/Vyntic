#!/usr/bin/env python3
"""Generate fictional Brightwater PE fund diligence and monitoring documents."""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    NextPageTemplate,
    PageBreak,
    PageTemplate,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from xml.sax.saxutils import escape


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"
SAMPLE_LINE = "SAMPLE - FICTIONAL, FOR DEMO USE"

LEDGER = {
    "gp": "Brightwater Capital Partners, LLC",
    "lp": "Glenmoor University Endowment",
    "fund_iii": "Brightwater Capital Partners III, L.P.",
    "fund_iv": "Brightwater Capital Partners IV, L.P.",
    "commitment": 25_000_000,
    "paid_in": 18_750_000,
    "distributed": 6_200_000,
    "nav": 21_400_000,
    "dpi": 0.33,
    "rvpi": 1.14,
    "tvpi": 1.47,
}

PARTNERS = [
    ("Elaine Hart", "Managing Partner", "Origination, Investment Committee Chair"),
    ("Malik Voss", "Partner", "Industrials and engineered products"),
    ("Priya Selvan", "Partner", "Business services and portfolio operations"),
    ("Daniel Roache", "Partner, Head of Value Creation", "Operational improvement and commercial excellence"),
]

TRACK_RECORD = [
    ["Brightwater Capital Partners I, L.P.", 2011, 300_000_000, 0.18, 2.10, 2.10, 0.00, 1.00, 0.00],
    ["Brightwater Capital Partners II, L.P.", 2015, 550_000_000, 0.21, 1.90, 1.30, 0.60, 0.68, 0.05],
    ["Brightwater Capital Partners III, L.P.", 2021, 850_000_000, 0.15, 1.50, 0.40, 0.95, 0.32, 0.08],
]


def money(value: int | float) -> str:
    return f"${value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1%}"


def date_text(value: dt.date) -> str:
    return value.strftime("%B %-d, %Y") if hasattr(value, "strftime") else str(value)


def make_styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="BWTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1f3f5b"),
            alignment=TA_CENTER,
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BWSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#4b5563"),
            spaceAfter=18,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Section",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            textColor=colors.HexColor("#1f3f5b"),
            spaceBefore=12,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Body",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.8,
            leading=11.5,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Small",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=7.4,
            leading=9,
            textColor=colors.HexColor("#374151"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="TableCell",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=7.6,
            leading=9.2,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SlideTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=29,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#17324d"),
            spaceAfter=14,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SlideKicker",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#6b7280"),
            spaceAfter=8,
        )
    )
    return styles


STYLES = make_styles()


def p(text: str, style: str = "Body") -> Paragraph:
    return Paragraph(escape(text), STYLES[style])


def rich(text: str, style: str = "Body") -> Paragraph:
    return Paragraph(text, STYLES[style])


def table(data: list[list[object]], widths: list[float] | None = None, font_size: float = 7.6) -> Table:
    converted = []
    for row in data:
        converted.append([
            cell if hasattr(cell, "wrap") else Paragraph(escape(str(cell)), STYLES["TableCell"])
            for cell in row
        ])
    t = Table(converted, colWidths=widths, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3f5b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("LEADING", (0, 0), (-1, -1), font_size + 1.8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return t


def draw_header_footer(canvas, doc, title: str):
    canvas.saveState()
    width, height = doc.pagesize
    canvas.setFont("Helvetica-Bold", 8)
    canvas.setFillColor(colors.HexColor("#1f3f5b"))
    canvas.drawString(0.55 * inch, height - 0.35 * inch, "Brightwater Capital Partners, LLC")
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#4b5563"))
    canvas.drawRightString(width - 0.55 * inch, height - 0.35 * inch, title[:70])
    canvas.setStrokeColor(colors.HexColor("#d1d5db"))
    canvas.line(0.55 * inch, height - 0.48 * inch, width - 0.55 * inch, height - 0.48 * inch)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawString(0.55 * inch, 0.35 * inch, SAMPLE_LINE)
    canvas.drawRightString(width - 0.55 * inch, 0.35 * inch, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf(filename: str, title: str, subtitle: str, elements: list, pagesize=LETTER):
    path = OUTPUT / filename
    doc = SimpleDocTemplate(
        str(path),
        pagesize=pagesize,
        rightMargin=0.62 * inch,
        leftMargin=0.62 * inch,
        topMargin=0.72 * inch,
        bottomMargin=0.62 * inch,
        pageCompression=0,
    )
    story = [
        p(title, "BWTitle"),
        p(subtitle, "BWSubtitle"),
        Spacer(1, 0.08 * inch),
    ]
    story.extend(elements)
    doc.build(story, onFirstPage=lambda c, d: draw_header_footer(c, d, title), onLaterPages=lambda c, d: draw_header_footer(c, d, title))


def section(num: str, title: str, paragraphs: Iterable[str], story: list):
    story.append(p(f"{num}. {title}", "Section"))
    for para in paragraphs:
        story.append(p(para))


def maybe_page_break(story: list):
    if story and not isinstance(story[-1], PageBreak):
        story.append(PageBreak())


def boilerplate(topic: str, count: int = 3) -> list[str]:
    base = [
        f"{topic} procedures are documented in internal memoranda, reviewed by finance, legal and compliance personnel where applicable, and retained in the Firm's books and records. The General Partner may modify these procedures in good faith to reflect market practice, investor requirements, legal developments or portfolio-company facts and circumstances.",
        f"The General Partner expects to apply judgment when implementing {topic.lower()} provisions. Such judgment may include consideration of timing, materiality, confidentiality, available information, administrative burden and whether the relevant matter is consistent with the Partnership's investment objectives.",
        f"No Limited Partner should assume that policies relating to {topic.lower()} eliminate all conflicts, valuation uncertainty, operational risk or timing differences. Remedies and approval rights are limited to those expressly set forth in the governing documents.",
        f"Records relating to {topic.lower()} are maintained by the Manager and may be summarized for the Advisory Committee, auditors, tax advisers, administrators or Limited Partners when the General Partner determines that disclosure is appropriate under the Partnership Agreement.",
    ]
    return base[:count]


def track_record_table():
    return table(
        [
            ["Fund", "Vintage", "Fund Size", "Net IRR", "TVPI", "DPI", "RVPI", "Status"],
            ["Fund I", "2011", "$300,000,000", "18%", "2.10x", "2.10x", "0.00x", "Fully realized"],
            ["Fund II", "2015", "$550,000,000", "21%", "1.90x", "1.30x", "0.60x", "Partially realized"],
            ["Fund III", "2021", "$850,000,000", "15%", "1.50x", "0.40x", "0.95x", "Active"],
        ],
        widths=[1.0 * inch, 0.6 * inch, 1.0 * inch, 0.7 * inch, 0.6 * inch, 0.6 * inch, 0.6 * inch, 1.2 * inch],
    )


def common_risk_paragraphs() -> list[str]:
    return [
        "An investment in the Partnership is speculative and involves a high degree of risk. Prospective investors should rely only on their own review of the Partnership Agreement, subscription materials, tax advice and diligence of the General Partner.",
        "Portfolio investments are expected to be illiquid, privately negotiated and subject to business, market, financing, operational, legal and regulatory risks. There can be no assurance that any portfolio company will achieve its business plan or that exit markets will be available on attractive terms.",
        "The General Partner and its affiliates may provide services to portfolio companies, receive transaction, directors, monitoring or similar fees, and allocate broken-deal, travel, consulting, insurance and other expenses among funds and accounts using policies that involve judgment.",
        "Performance information regarding prior funds is not necessarily indicative of future results. Valuations of unrealized investments are inherently uncertain and may differ materially from values that would be realized in an orderly sale.",
    ]


def generate_ppm(run_date: dt.date):
    story: list = []
    section("1", "Notice to Prospective Investors", [
        "This private placement memorandum relates to the offering of limited partnership interests in Brightwater Capital Partners IV, L.P., a Delaware limited partnership formed to pursue control-oriented investments in North American industrial and business services companies. The offering is made only to eligible investors on a confidential basis.",
        "All names, amounts and entities in this document are fictional and are provided for demonstration use. No public offering is being made and no securities are being sold.",
        *boilerplate("Offering restrictions", 2),
    ], story)
    maybe_page_break(story)
    section("2", "Executive Summary", [
        "Brightwater Capital Partners, LLC is a Chicago-based middle-market buyout sponsor founded in 2009 with approximately $2.1 billion of assets under management. The Firm targets durable industrial niches and outsourced business services where operating discipline, pricing analytics and professionalized sales execution can create value.",
        "Fund IV is targeting $1.25 billion of commitments, with a hard cap of $1.5 billion. The General Partner expects to invest in 10 to 14 platform companies with enterprise values generally between $75 million and $350 million.",
        *boilerplate("Fund IV investment program", 2),
    ], story)
    story.append(table([
        ["Term", "Summary"],
        ["Fund", "Brightwater Capital Partners IV, L.P."],
        ["Target / hard cap", "$1.25 billion target; $1.5 billion hard cap"],
        ["Management fee", "2.0% on commitments during the five-year investment period; thereafter 1.5% on invested capital"],
        ["Carried interest", "20%, European whole-fund waterfall, 8% preferred return, 100% GP catch-up and clawback"],
        ["GP commitment", "2.0% of aggregate commitments"],
        ["Fee offset", "50% of transaction, monitoring, directors and similar fees offset against the management fee"],
        ["Recycling / expenses", "Recycling permitted for 24 months; organizational expense cap of $2,500,000"],
        ["Removal", "No-fault removal of the General Partner requires approval of Limited Partners holding 80% of interests"],
    ], widths=[1.6 * inch, 4.8 * inch]))
    maybe_page_break(story)
    section("3", "Investment Strategy", [
        "The Fund will seek control investments in North American companies with resilient demand, recurring aftermarket exposure, fragmented customer bases and identifiable operational improvement opportunities. Brightwater expects to emphasize businesses serving safety, compliance, industrial automation, facility services, specialty distribution and technical field services markets.",
        "Brightwater typically partners with founders, family-owned businesses and corporate carve-outs where professionalized reporting, working-capital discipline, procurement programs and add-on acquisition integration can improve EBITDA margins. Fund IV will avoid early-stage technology, commodity cyclicality and businesses whose primary value depends on uncontracted government reimbursement.",
        "The investment committee process requires a preliminary memorandum, confirmatory diligence plan, quality of earnings review, debt-capacity analysis, legal and tax structuring review, and final investment committee approval. The Firm expects to use moderate leverage with conservative covenant and liquidity assumptions.",
        *boilerplate("Investment strategy", 2),
    ], story)
    maybe_page_break(story)
    section("4", "Senior Team", [
        "Brightwater is led by four partners supported by twelve investment professionals and a seven-person operating resource group. The Firm describes the senior team as stable and cohesive, with shared economics across the investment committee.",
        *boilerplate("Senior team oversight", 2),
    ], story)
    story.append(table([["Name", "Title", "Current responsibilities"]] + [[a, b, c] for a, b, c in PARTNERS], widths=[1.35 * inch, 1.6 * inch, 3.35 * inch]))
    bios = [
        "Elaine Hart founded Brightwater after senior investing roles at fictional regional sponsors and chairs the investment committee for Fund IV.",
        "Malik Voss leads industrial products coverage and has served on the boards of Amerin Gauge Systems, Caldera Flow Controls and Northline Fabrication, each fictional portfolio companies.",
        "Priya Selvan leads business services coverage and has overseen sourcing in facility compliance, testing services and specialty distribution.",
        "Daniel Roache serves as Partner and Head of Value Creation, with responsibility for operating diagnostics, management dashboards, procurement initiatives and pricing programs across the portfolio.",
    ]
    for bio in bios:
        story.append(p(bio))
    maybe_page_break(story)
    section("5", "Prior Fund Performance", [
        "The following summary is presented on a net basis and should be read with the detailed track-record schedule furnished separately. Unrealized values are determined under the Firm's valuation policy and are subject to change.",
        *boilerplate("Track record presentation", 2),
    ], story)
    story.append(track_record_table())
    maybe_page_break(story)
    for i in range(6, 11):
        section(str(i), ["Portfolio Construction", "Diligence Process", "Valuation Framework", "Conflicts and Fees", "Tax, ERISA and Regulatory Matters"][i - 6], [
            "Brightwater will construct the portfolio over the investment period with concentration limits, industry guidelines and reserve planning reviewed quarterly by the investment committee. Follow-on investments may be made to protect value, fund add-on acquisitions or support working capital during periods of growth.",
            "The Firm's diligence workstreams customarily include commercial market interviews, customer cohort analysis, purchase accounting review, environmental assessment, benefits review, insurance mapping and information-security diligence. Findings are summarized in a final investment memorandum.",
            *boilerplate(["Portfolio construction", "Diligence process", "Valuation framework", "Conflicts and fees", "Tax and regulatory matters"][i - 6], 3),
        ], story)
        maybe_page_break(story)
    section("11", "Risk Factors", common_risk_paragraphs() * 3, story)
    maybe_page_break(story)
    section("12", "Subscription Procedures", [
        "Prospective investors must complete subscription materials, provide beneficial ownership information and satisfy anti-money-laundering review before admission. The General Partner may reject any subscription in whole or in part.",
        f"This memorandum is dated {date_text(run_date)} and supersedes prior discussion materials relating to Fund IV.",
        *boilerplate("Subscription procedures", 2),
    ], story)
    build_pdf("brightwater_iv_ppm.pdf", "Brightwater Capital Partners IV, L.P.", "Private Placement Memorandum | Vyntic token: ppm", story)


def slide_page(title: str, bullets: list[str], footer: str | None = None) -> list:
    elements: list = [p("CONFIDENTIAL FUND IV MARKETING MATERIAL", "SlideKicker"), p(title, "SlideTitle")]
    for bullet in bullets:
        elements.append(rich(f"<bullet>&bull;</bullet> {escape(bullet)}", "Body"))
        elements.append(Spacer(1, 0.04 * inch))
    if footer:
        elements.append(Spacer(1, 0.15 * inch))
        elements.append(p(footer, "Small"))
    elements.append(PageBreak())
    return elements


def generate_pitchbook():
    story: list = []
    slides = [
        ("Top-quartile track record, 2.1x net", [
            "Brightwater Capital Partners is a Chicago-based middle-market buyout platform focused on North American industrials and business services.",
            "Fund IV is targeting $1.25 billion, with a $1.5 billion hard cap, to continue a control-oriented strategy refined since 2009.",
            "The Firm highlights a 2.1x net multiple from realized Fund I and a consistent underwriting process across three predecessor funds.",
        ], "The 2.1x net claim reflects Fund I, which is fully realized."),
        ("Where Brightwater invests", [
            "Industrial technology, compliance services, engineered products, specialty distribution and outsourced facility services.",
            "Typical entry enterprise values of $75 million to $350 million with identifiable add-on acquisition pathways.",
            "North American orientation with limited currency and cross-border complexity.",
        ], None),
        ("Prior fund snapshot", [
            "Fund I: $300 million, 2011 vintage, 18% net IRR, 2.10x net TVPI and 2.10x DPI.",
            "Fund II: $550 million, 2015 vintage, 21% net IRR, 1.90x net TVPI, 1.30x DPI and 0.60x RVPI.",
            "Fund III: $850 million, 2021 vintage, 15% net IRR and 1.50x net TVPI as presented in the marketing track record.",
        ], None),
        ("Senior team", [
            "Elaine Hart, Managing Partner - investment committee chair.",
            "Malik Voss, Partner - industrials lead.",
            "Priya Selvan, Partner - business services lead.",
            "Daniel Roache, Partner, Head of Value Creation - operating improvement programs and dashboard discipline.",
        ], None),
        ("Investment process", [
            "Dedicated sourcing by vertical, preliminary investment committee review and confirmatory diligence with independent quality of earnings support.",
            "Final investment committee approval requires unanimous approval of voting partner members present at a duly called meeting.",
            "Post-close value creation plan established within 100 days with KPI baseline, pricing workplan and acquisition roadmap.",
        ], None),
        ("Fund IV terms", [
            "2.0% management fee on commitments during the investment period; 1.5% on invested capital thereafter.",
            "20% carried interest, 8% preferred return, European waterfall, GP catch-up and clawback.",
            "2.0% GP commitment and customary advisory committee oversight.",
        ], None),
        ("Value creation examples", [
            "Pricing analytics and quote governance in engineered products.",
            "Procurement and working-capital discipline in specialty distribution.",
            "Add-on integration playbooks for route density and branch-level margin improvement.",
        ], None),
        ("Portfolio construction", [
            "10 to 14 platform investments, generally with no single initial investment exceeding 15% of commitments without advisory committee consent.",
            "Reserve model reviewed quarterly for each platform and add-on acquisition pathway.",
            "Expected holding periods of four to six years.",
        ], None),
        ("Reporting and investor relations", [
            "Quarterly reports, annual audited financial statements and annual investor meeting materials.",
            "Capital account statements provided to limited partners following each quarter-end close.",
            "Data room includes PPM, LPA, DDQ, track record, valuation policy and compliance brochure.",
        ], None),
        ("Why Fund IV", [
            "Continuity of middle-market strategy and sector focus.",
            "Operating resources aligned to pricing, procurement and commercial process improvement.",
            "Fund III presented at 1.50x net TVPI in current marketing materials.",
        ], "Past performance is not indicative of future results."),
    ]
    for title, bullets, foot in slides:
        story.extend(slide_page(title, bullets, foot))
    if story and isinstance(story[-1], PageBreak):
        story.pop()
    build_pdf("brightwater_iv_pitchbook.pdf", "Brightwater Capital Partners IV, L.P.", "Marketing Presentation | Vyntic token: pitchbook", story, pagesize=landscape(LETTER))


def generate_track_record_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Track Record"
    ws["A1"] = "Brightwater Capital Partners Track Record"
    ws["A2"] = SAMPLE_LINE
    ws["A3"] = "All names and entities are fictional. Values are net to limited partners and shown as of the marketing track-record cut-off."
    headers = ["Fund", "Vintage", "Fund Size", "Net IRR", "TVPI", "DPI", "RVPI", "Realized%", "Loss Ratio"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)
    for r, row in enumerate(TRACK_RECORD, 6):
        for c, value in enumerate(row, 1):
            ws.cell(row=r, column=c, value=value)
    widths = [38, 10, 16, 11, 10, 10, 10, 12, 12]
    for c, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = width
    dark = PatternFill("solid", fgColor="1F3F5B")
    pale = PatternFill("solid", fgColor="EAF2F8")
    thin = Side(style="thin", color="CBD5E1")
    for row in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    ws["A1"].font = Font(bold=True, size=16, color="1F3F5B")
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A3"].font = Font(size=9, color="666666")
    for cell in ws[5]:
        cell.fill = dark
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=6, max_row=8, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if cell.row % 2 == 0:
                cell.fill = pale
    for r in range(6, 9):
        ws.cell(r, 3).number_format = '"$"#,##0'
        ws.cell(r, 4).number_format = "0%"
        ws.cell(r, 5).number_format = "0.00x"
        ws.cell(r, 6).number_format = "0.00x"
        ws.cell(r, 7).number_format = "0.00x"
        ws.cell(r, 8).number_format = "0%"
        ws.cell(r, 9).number_format = "0%"
    ws.freeze_panes = "A6"
    wb.save(OUTPUT / "brightwater_track_record.xlsx")


def generate_ddq(run_date: dt.date):
    qa = [
        ("1. Firm & Ownership", "Describe the Firm, ownership and assets under management.", "Brightwater Capital Partners, LLC was founded in 2009 and is headquartered in Chicago. The Firm manages approximately $2.1 billion across three predecessor buyout funds and related co-investment vehicles. Ownership is held by the senior investment partners and a small number of non-voting legacy members."),
        ("2. Team & Succession", "Describe senior team stability, succession planning and key-person coverage.", "The Firm believes it has an institutionalized team model with shared investment committee participation, cross-staffed deal teams and written succession protocols. Senior responsibility for sourcing, execution and portfolio value creation is allocated across multiple partners and committees, and the Firm expects continuity of coverage throughout Fund IV's investment period."),
        ("3. Investment Committee", "Who approves investments?", "Elaine Hart chairs the Investment Committee, with voting participation from Malik Voss, Priya Selvan and Daniel Roache. The committee reviews preliminary, confirmatory and final memoranda before capital is committed."),
        ("4. Track Record", "Summarize predecessor fund performance.", "Fund I is fully realized at 2.10x TVPI and 18% net IRR. Fund II is marked at 1.90x TVPI and 21% net IRR. Fund III is presented in the marketing track record at 1.50x TVPI and 15% net IRR."),
        ("5. Strategy & Process", "What is the Fund IV investment strategy?", "Fund IV will continue Brightwater's focus on control investments in North American industrials and business services. The Firm targets companies with recurring aftermarket demand, fragmented customer sets and opportunities for pricing, procurement and working-capital improvement."),
        ("6. Fund Terms", "Summarize the proposed economics.", "The Firm views the terms as market standard for a sponsor of Brightwater's size and history. Management fees, carried interest, organizational expenses and governance rights are documented in the draft Partnership Agreement."),
        ("7. Fees and Offsets", "Describe management fees and fee offsets.", "Management fee arrangements are market standard, and the Fund provides a 100% fee offset for transaction, monitoring, directors and similar fees received by the General Partner or its affiliates."),
        ("8. Valuation", "Describe valuation governance.", "Portfolio investments are valued quarterly under the Firm's written valuation policy. The valuation committee reviews company performance, market comparables, leverage, liquidity and transaction evidence, with third-party review incorporated annually for Level 3 assets."),
        ("9. Compliance & Regulatory", "Describe compliance resources.", "The Firm maintains a compliance manual, code of ethics, restricted-list process, annual compliance review and employee certification procedures. Regulatory correspondence is managed by the chief compliance officer and outside counsel."),
        ("10. IT & Cybersecurity", "Describe cybersecurity controls.", "Brightwater uses reputable cloud-hosted systems, multi-factor authentication, endpoint protection and employee awareness reminders. The Firm periodically discusses cyber readiness with outside technology vendors and updates policies when needed. A formal SOC 2 report is not maintained, and testing cadence is determined by management."),
        ("11. ESG", "Describe responsible investment approach.", "The Firm incorporates material environmental, labor, safety and governance considerations into diligence where relevant to the target business and provides LP communications on portfolio initiatives."),
        ("12. Conflicts", "Identify affiliated service providers and conflicts.", "Potential conflicts are addressed through the Firm's compliance manual, advisory committee process and allocation policy. The Firm does not expect Fund IV to rely on affiliated service providers in the ordinary course other than the General Partner and management company."),
        ("13. Allocation", "How are co-investment and broken-deal expenses allocated?", "Co-investments are allocated based on capacity, strategic fit, timing and legal suitability. Broken-deal expenses are allocated among participating vehicles based on the opportunity pursued and benefits expected."),
        ("14. Service Providers", "List core service providers.", "Fund administration is expected to be performed by North Pier Fund Services. Audit services are expected to be performed by Huxley Markham & Co. LLP, and fund counsel is expected to be Alder & Finch LLP."),
        ("15. Use of Leverage", "Describe leverage policy.", "The Firm expects moderate company-level leverage and may use subscription facilities for short-term working capital, bridging capital calls and administrative efficiency."),
        ("16. Portfolio Monitoring", "Describe monitoring cadence.", "Portfolio company boards meet at least quarterly. The Firm maintains monthly operating dashboards for revenue, EBITDA, working capital, liquidity, safety and acquisition pipeline metrics."),
        ("17. Reporting", "Describe LP reporting.", "Limited partners receive quarterly reports, capital account statements, capital call notices, distribution notices and audited annual financial statements according to governing documents."),
        ("18. References", "Provide reference approach.", "References are available from portfolio executives, lenders, co-investors and limited partners subject to confidentiality and scheduling constraints."),
        ("19. Side Letters", "How are side-letter obligations tracked?", "Side-letter terms are reviewed by legal, finance and investor relations personnel and maintained in an internal obligations tracker."),
        ("20. Closing Timetable", "Describe expected closing timetable.", f"Fund IV expects to hold closings during 2026, subject to investor diligence and documentation. This DDQ response is dated {date_text(run_date)}."),
    ]
    story: list = []
    for heading, question, answer in qa:
        story.append(p(heading, "Section"))
        story.append(rich(f"<b>Question:</b> {escape(question)}", "Body"))
        story.append(rich(f"<b>Response:</b> {escape(answer)}", "Body"))
        for extra in boilerplate(heading, 2):
            story.append(p(extra))
        if heading != qa[-1][0]:
            maybe_page_break(story)
    build_pdf("brightwater_iv_ddq.pdf", "Brightwater Capital Partners IV, L.P.", "Completed ILPA-style DDQ | Vyntic token: ddq", story)


def generate_lpa():
    clauses = [
        ("1", "Formation; Name; Term", "The Partnership is formed as a Delaware limited partnership under the name Brightwater Capital Partners IV, L.P. The Partnership shall continue until the tenth anniversary of the final closing, subject to two one-year extensions approved by the Advisory Committee."),
        ("2", "Purpose", "The purpose of the Partnership is to make, hold, manage and dispose of control-oriented investments in North American industrial and business services companies and to engage in related activities."),
        ("3", "Capital Commitments", "The target aggregate commitments are $1,250,000,000, and the General Partner may accept commitments up to a hard cap of $1,500,000,000. The General Partner and its affiliates shall commit not less than 2.0% of aggregate commitments."),
        ("4", "Investment Period", "The investment period shall begin on the initial closing and end on the fifth anniversary thereof, unless earlier terminated under this Agreement."),
        ("5", "Management Fee", "During the investment period, the Partnership shall pay a management fee equal to 2.0% per annum of aggregate commitments. Thereafter, the fee shall equal 1.5% per annum of invested capital, calculated quarterly in advance."),
        ("6", "Fee Offset", "Fifty percent (50%) of transaction fees, monitoring fees, directors fees, break-up fees and similar fees received by the General Partner, Manager or their affiliates from portfolio companies shall offset the management fee otherwise payable by the Partnership."),
        ("7", "Distributions; Waterfall", "Distributions shall be made on a whole-fund European waterfall. Limited Partners shall first receive return of contributed capital and an 8% preferred return, followed by a 100% catch-up to the General Partner until the General Partner has received 20% of aggregate profits, and thereafter 80% to Limited Partners and 20% carried interest to the General Partner."),
        ("8", "Clawback", "The General Partner shall be subject to a customary after-tax clawback intended to return excess carried interest following liquidation of the Partnership."),
        ("9", "Recycling", "The General Partner may recall and recycle distributions representing cost basis, bridge financing proceeds and certain other amounts for a period of 24 months after the original investment date, subject to the limitations set forth herein."),
        ("10", "Organizational Expenses", "The Partnership shall bear organizational and offering expenses up to an aggregate cap of $2,500,000. Expenses in excess of such cap shall be borne by the Manager or its affiliates."),
        ("11", "Key Persons", "The Key Persons are Elaine Hart and Daniel Roache. A Key Person Event shall occur if fewer than two Key Persons devote substantially all of their business time to the affairs of the Partnership and related Brightwater funds, subject to cure and Limited Partner Advisory Committee waiver rights."),
        ("12", "No-Fault Removal", "The General Partner may be removed without cause only upon approval of Limited Partners holding at least 80% of aggregate commitments, excluding interests held by the General Partner and its affiliates."),
        ("13", "Advisory Committee", "The Advisory Committee shall review conflicts, valuation matters referred by the General Partner, investment-period extensions, certain affiliate transactions and other matters set forth in this Agreement."),
        ("14", "Excuse and Exclusion", "A Limited Partner may be excused from a particular investment where participation would reasonably be expected to violate law, regulation or written policy accepted by the General Partner."),
        ("15", "Transfers", "No Limited Partner may transfer interests without the prior written consent of the General Partner, which may be withheld in accordance with this Agreement and applicable law."),
        ("16", "Books; Reports", "The Partnership shall maintain books and records and provide quarterly unaudited reports, annual audited financial statements and tax information as reasonably determined by the General Partner."),
        ("17", "Indemnification", "The Partnership shall indemnify the General Partner, Manager and covered persons to the fullest extent permitted by law for actions taken in good faith on behalf of the Partnership."),
        ("18", "Confidentiality", "Limited Partners shall keep confidential all non-public information regarding the Partnership, portfolio companies and other Limited Partners, subject to customary exceptions."),
        ("19", "Amendments", "This Agreement may be amended by the General Partner with the consent of Limited Partners holding a majority in interest, except for amendments requiring a higher threshold or adversely affecting a specific Limited Partner disproportionately."),
        ("20", "Governing Law", "This Agreement shall be governed by the laws of the State of Delaware."),
    ]
    story: list = []
    for number, title, body in clauses:
        section(number, title, [
            body,
            "The provisions of this Section shall be interpreted consistently with the definitions, limitations and approval mechanics set forth elsewhere in this Agreement.",
            *boilerplate(title, 2),
        ], story)
        if number != clauses[-1][0]:
            maybe_page_break(story)
    build_pdf("brightwater_iv_lpa.pdf", "Brightwater Capital Partners IV, L.P.", "Limited Partnership Agreement | Vyntic token: lpa", story)


def generate_adv(run_date: dt.date):
    items = [
        ("Item 1. Cover Page", f"Brightwater Capital Partners, LLC is a fictional registered investment adviser headquartered in Chicago. This brochure is dated {date_text(run_date)} and is provided solely for demonstration use."),
        ("Item 4. Advisory Business", "Brightwater provides discretionary private fund management services to buyout funds and related co-investment vehicles. The Firm manages approximately $2.1 billion in assets. Daniel Roache ceased being an advisory employee of the Firm effective February 28, 2026, following a transition of portfolio operations responsibilities to the remaining partner group and operating resources."),
        ("Item 5. Fees and Compensation", "The private funds generally pay management fees and carried interest as set forth in their governing documents. Portfolio companies may pay transaction, monitoring, directors or similar fees to Brightwater or affiliates, subject to offsets and allocations described in the applicable fund documents."),
        ("Item 6. Performance-Based Fees", "Brightwater and its affiliates may receive performance-based carried interest. Such arrangements may create incentives to make or hold investments with greater risk or return profiles."),
        ("Item 8. Methods of Analysis", "Brightwater uses fundamental company diligence, industry research, financial modeling, third-party accounting and legal workstreams, management interviews and operational diagnostics."),
        ("Item 10. Other Financial Industry Activities and Affiliations", "Brightwater Securities, LLC is an affiliated broker-dealer under common control with the Manager. Brightwater Securities, LLC may receive transaction fees or placement-related compensation in connection with portfolio company transactions, subject to disclosure, allocation and applicable law."),
        ("Item 11. Code of Ethics", "The Firm maintains a code of ethics, personal trading policy, restricted list and annual certification process for supervised persons."),
        ("Item 12. Brokerage Practices", "The Firm generally does not direct brokerage for public securities trading, but may recommend transaction counterparties, lenders, consultants and other vendors in private-company transactions."),
        ("Item 14. Client Referrals and Other Compensation", "The Firm and affiliates may receive transaction, monitoring, directors or similar compensation from portfolio companies. Such compensation may create conflicts that are addressed through fund documents, advisory committee review where applicable and compliance policies."),
        ("Item 18. Financial Information", "In 2023, the Firm received an SEC deficiency letter concerning documentation of expense allocation among funds and co-investment vehicles. The Firm enhanced written allocation procedures, remediated sampled allocations and completed additional employee training. No monetary penalty was imposed."),
    ]
    story: list = []
    for heading, body in items:
        section(heading.split(".")[0].replace("Item ", ""), heading, [
            body,
            "Additional information is available upon request, subject to confidentiality and eligibility requirements.",
            *boilerplate(heading, 2),
        ], story)
        if heading != items[-1][0]:
            maybe_page_break(story)
    build_pdf("brightwater_adv_part2a.pdf", "Brightwater Capital Partners, LLC", "Form ADV Part 2A Brochure | Vyntic token: form_adv", story)


def generate_valuation_policy():
    story: list = []
    sections = [
        ("1", "Purpose and Scope", "This policy establishes procedures for valuing portfolio investments held by Brightwater-sponsored private funds, including Level 3 privately held investments for which observable market prices are not available."),
        ("2", "Governance", "The valuation committee meets quarterly and includes finance, investment, compliance and portfolio operations representatives. The committee reviews valuation memoranda, company performance, market data, financing terms, exit assumptions and material events."),
        ("3", "Methodologies", "Valuations may use comparable company multiples, precedent transactions, discounted cash flow analyses, recent financing transactions, cost, indications of interest and other methods deemed appropriate in light of facts and circumstances."),
        ("4", "Level 3 Review", "Level 3 portfolio company marks are prepared by the investment team and reviewed by finance and the valuation committee each quarter. Third-party valuation review is obtained annually for active Level 3 assets and may also be obtained for material events or advisory committee requests."),
        ("5", "Documentation", "Each quarterly valuation file shall include management reporting, debt schedules, working-capital trends, market multiples, calibration to transaction price, selected multiple support and committee approvals."),
        ("6", "Conflicts and Escalation", "Disagreements are escalated to the chief financial officer and chief compliance officer. Material methodology changes, conflicts or valuation exceptions may be reported to the Limited Partner Advisory Committee."),
    ]
    for num, title, body in sections:
        section(num, title, [
            body,
            "The policy is intended to promote consistency and auditability while recognizing that private-company valuations require judgment.",
            *boilerplate(title, 2),
        ], story)
        if num != sections[-1][0]:
            maybe_page_break(story)
    build_pdf("brightwater_valuation_policy.pdf", "Brightwater Capital Partners, LLC", "Valuation Policy | Vyntic token: valuation_policy", story)


def generate_side_letter():
    obligations = [
        ("1", "Management Fee Reduction", "With respect to Glenmoor University Endowment's $25,000,000 commitment to Brightwater Capital Partners III, L.P., the management fee otherwise applicable to Glenmoor shall be reduced by ten basis points per annum."),
        ("2", "Most Favored Nations", "If the General Partner grants a better economic, reporting, transfer or governance term to any Limited Partner committing $50,000,000 or less, Glenmoor shall have an MFN election right exercisable within 30 days after final close, subject to customary exclusions."),
        ("3", "Co-Investment Opportunity", "For any portfolio investment requiring more than $75,000,000 of aggregate equity capital from the Fund and affiliated vehicles, the General Partner shall use commercially reasonable efforts to offer Glenmoor a pro-rata co-investment opportunity."),
        ("4", "Reporting", "The General Partner shall provide quarterly reports within 45 days after quarter-end and annual audited financial statements within 120 days after fiscal year-end."),
        ("5", "Excuse Right", "Upon at least ten business days' written notice, Glenmoor may be excused from participation in investments whose primary business is tobacco or controlled substances."),
        ("6", "Transfers to Affiliates", "The General Partner's consent to transfers by Glenmoor to Glenmoor affiliates shall not be unreasonably withheld, conditioned or delayed, subject to legal, tax and regulatory limitations."),
        ("7", "Annual ESG Report", "The General Partner shall provide an annual ESG report covering portfolio-level metrics reasonably available to the General Partner."),
        ("8", "Confidentiality; No Other Modification", "Except as expressly set forth herein, the Partnership Agreement remains unchanged. This letter agreement is confidential and may be disclosed only as permitted by the Partnership Agreement or applicable law."),
    ]
    story: list = []
    story.append(p("This side letter is entered into by Brightwater Capital Partners III GP, LLC, as general partner of Brightwater Capital Partners III, L.P., and Glenmoor University Endowment.", "Body"))
    for num, title, body in obligations:
        section(num, title, [body, *boilerplate(title, 1)], story)
        if num in {"2", "4", "6"}:
            maybe_page_break(story)
    build_pdf("glenmoor_fund_iii_side_letter.pdf", "Brightwater Capital Partners III, L.P.", "Side Letter with Glenmoor University Endowment | Vyntic token: side_letter", story)


def generate_pcap():
    story: list = []
    section("1", "Account Summary", [
        "The following partners' capital account statement is provided to Glenmoor University Endowment for Brightwater Capital Partners III, L.P. as of June 30, 2026.",
    ], story)
    story.append(table([
        ["Metric", "Amount"],
        ["Commitment", "$25,000,000"],
        ["Paid-in / called capital", "$18,750,000"],
        ["Percent called", "75.0%"],
        ["Cumulative distributions", "$6,200,000"],
        ["NAV / remaining value", "$21,400,000"],
        ["DPI", "0.33x"],
        ["RVPI", "1.14x"],
        ["TVPI", "1.47x"],
    ], widths=[2.4 * inch, 2.0 * inch]))
    story.append(p("DPI plus RVPI equals TVPI based on rounded values: 0.33x + 1.14x = 1.47x.", "Small"))
    section("2", "Quarterly Capital Activity", [
        "The management fee allocation reflects Glenmoor's side-letter fee rate of 1.90% per annum, equal to the standard 2.00% rate less a ten basis point annual reduction.",
    ], story)
    story.append(table([
        ["Q2 2026 Activity", "Amount"],
        ["Beginning capital account", "$20,900,000"],
        ["Capital contributions", "$0"],
        ["Distributions", "$0"],
        ["Management fee allocation", "($118,750)"],
        ["Partnership expense allocation", "($66,250)"],
        ["Carried interest accrual", "($315,000)"],
        ["Net unrealized appreciation", "$1,000,000"],
        ["Ending capital account / NAV", "$21,400,000"],
    ], widths=[2.8 * inch, 1.8 * inch]))
    build_pdf("glenmoor_fund_iii_pcap_q2_2026.pdf", "Brightwater Capital Partners III, L.P.", "Glenmoor Partners' Capital Account Statement - Q2 2026 | Vyntic token: capital_account", story)


def generate_quarterly():
    story: list = []
    report_date = dt.date(2026, 8, 29)
    section("1", "GP Letter", [
        f"This quarterly report for Brightwater Capital Partners III, L.P. is dated {date_text(report_date)} and covers the quarter ended June 30, 2026.",
        "Fund III continued to focus on portfolio operating initiatives, add-on acquisition sourcing and liquidity management during the quarter. Public-market multiple contraction modestly affected comparable-company marks, while company-level EBITDA growth offset part of the impact.",
        *boilerplate("Quarterly portfolio reporting", 3),
    ], story)
    maybe_page_break(story)
    section("2", "Fund-Level Performance", [
        "Fund III is an $850,000,000 2021 vintage buyout fund. As of June 30, 2026, Glenmoor University Endowment's capital account reflected a $25,000,000 commitment, $18,750,000 of paid-in capital, $6,200,000 of cumulative distributions and $21,400,000 of NAV.",
        *boilerplate("Fund-level performance reporting", 2),
    ], story)
    story.append(table([
        ["Metric", "Glenmoor Q2 2026"],
        ["Commitment", "$25,000,000"],
        ["Paid-in / called", "$18,750,000"],
        ["Distributed", "$6,200,000"],
        ["NAV / remaining value", "$21,400,000"],
        ["DPI", "0.33x"],
        ["RVPI", "1.14x"],
        ["TVPI", "1.47x"],
    ], widths=[2.5 * inch, 2.0 * inch]))
    maybe_page_break(story)
    companies = [
        ("Amerin Gauge Systems", "Engineered measurement components", "$132.0 million", "Revenue increased 6.4% year-over-year, with margin expansion from procurement savings and mix improvement."),
        ("Caldera Flow Controls", "Specialty valves and flow control", "$188.5 million", "Management completed a plant consolidation plan and continues to pursue add-on opportunities in aftermarket service."),
        ("Northline Fabrication", "Precision metal fabrication", "$96.0 million", "Order backlog remains above underwriting case, though customer concentration is being monitored."),
        ("Vantage Field Services", "Facility compliance services", "$165.5 million", "Route density and technician utilization improved during the quarter following dispatch software rollout."),
    ]
    section("3", "Portfolio Company Updates", ["Portfolio updates are summarized below based on unaudited company reporting and Brightwater valuation committee review."], story)
    for company in companies:
        story.append(table([["Company", "Sector", "Fund III fair value", "Quarterly update"], company], widths=[1.35 * inch, 1.35 * inch, 1.05 * inch, 2.65 * inch]))
        for extra in boilerplate(company[0], 3):
            story.append(p(extra))
        maybe_page_break(story)
    section("4", "Capital Account and Fee Disclosure", [
        "Glenmoor's Q2 management fee allocation was $118,750, reflecting the 1.90% annual management fee rate applicable under Glenmoor's side letter. The standard rate for similarly situated investors before the discount is 2.00% per annum.",
        "No capital contributions or distributions were posted to Glenmoor's capital account during the second quarter. Post-quarter activity will be reported in subsequent notices and capital account statements.",
        *boilerplate("Capital account and fee disclosure", 2),
    ], story)
    build_pdf("brightwater_iii_quarterly_q2_2026.pdf", "Brightwater Capital Partners III, L.P.", "Quarterly Report - Quarter Ended June 30, 2026 | Vyntic token: quarterly_report", story)


def generate_capital_call(run_date: dt.date):
    due = run_date + dt.timedelta(days=5)
    story: list = []
    section("1", "Notice", [
        f"Capital Call Notice No. 7 is issued to Glenmoor University Endowment on {date_text(run_date)} with payment due no later than {date_text(due)}.",
        "The capital call amount is $1,875,000, representing 7.5% of Glenmoor's $25,000,000 commitment to Brightwater Capital Partners III, L.P.",
    ], story)
    story.append(table([
        ["Item", "Amount / Detail"],
        ["Investor", "Glenmoor University Endowment"],
        ["Commitment", "$25,000,000"],
        ["Call percentage", "7.5%"],
        ["Capital call amount", "$1,875,000"],
        ["Purpose", "New platform acquisition: Project Cardinal"],
        ["Remaining unfunded commitment after this call", "$4,375,000"],
        ["Wire bank", "Anchorstone Bank, N.A. (fictional)"],
        ["Account name", "Brightwater Capital Partners III, L.P. Capital Account"],
        ["ABA / account", "000000000 / 000123456789"],
    ], widths=[2.5 * inch, 3.3 * inch]))
    build_pdf("brightwater_iii_capital_call_07.pdf", "Brightwater Capital Partners III, L.P.", "Capital Call Notice No. 7 | Vyntic token: capital_call", story)


def generate_distribution(run_date: dt.date):
    pay = run_date + dt.timedelta(days=12)
    story: list = []
    section("1", "Notice", [
        f"Distribution Notice No. 3 is issued to Glenmoor University Endowment on {date_text(run_date)}. Payment is expected on {date_text(pay)}, subject to banking cut-off times.",
        "The distribution relates to partial proceeds from a dividend recapitalization and working-capital release at a Fund III portfolio company.",
    ], story)
    story.append(table([
        ["Distribution Component", "Amount"],
        ["Return of capital", "$950,000"],
        ["Realized gain", "$450,000"],
        ["Total distribution", "$1,400,000"],
    ], widths=[2.6 * inch, 1.6 * inch]))
    story.append(table([
        ["Investor", "Glenmoor University Endowment"],
        ["Payment date", date_text(pay)],
        ["Payment method", "Wire transfer to bank instructions on file"],
        ["Tax note", "Tax reporting will be provided with annual Schedule K-1 materials."],
    ], widths=[2.2 * inch, 3.4 * inch]))
    build_pdf("brightwater_iii_distribution_03.pdf", "Brightwater Capital Partners III, L.P.", "Distribution Notice No. 3 | Vyntic token: distribution_notice", story)


def generate_financial_statements():
    story: list = []
    section("1", "Independent Auditor's Report", [
        "Huxley Markham & Co. LLP, independent auditors, have audited the accompanying fictional financial statements of Brightwater Capital Partners III, L.P. as of and for the year ended December 31, 2025. In our opinion, the financial statements present fairly, in all material respects, the financial position of the Partnership in accordance with accounting principles generally accepted in the United States of America.",
        *boilerplate("Audit opinion", 3),
    ], story)
    maybe_page_break(story)
    section("2", "Statement of Assets and Liabilities", [
        "The Partnership is an $850,000,000 2021 vintage private equity fund. The figures below are presented in thousands except where noted and are fictional.",
        *boilerplate("Statement of assets and liabilities", 2),
    ], story)
    story.append(table([
        ["Assets / Liabilities", "December 31, 2025"],
        ["Investments, at fair value", "$610,000"],
        ["Cash and restricted cash", "$18,750"],
        ["Receivables and other assets", "$7,250"],
        ["Total assets", "$636,000"],
        ["Subscription facility payable", "($21,000)"],
        ["Accrued expenses and management fees", "($5,000)"],
        ["Partners' capital / NAV", "$610,000"],
    ], widths=[3.0 * inch, 2.0 * inch]))
    maybe_page_break(story)
    section("3", "Statement of Operations", [
        "For the year ended December 31, 2025, the Partnership recorded investment income of $8.5 million, management fees of $17.0 million, partnership expenses of $4.1 million, realized gains of $42.0 million and net unrealized appreciation of $88.6 million.",
        *boilerplate("Statement of operations", 3),
    ], story)
    maybe_page_break(story)
    section("4", "Schedule of Investments", [
        "The following portfolio-company schedule is summarized and excludes immaterial escrow, cash and tax-basis details.",
        *boilerplate("Schedule of investments", 2),
    ], story)
    story.append(table([
        ["Portfolio Company", "Sector", "Cost", "Fair Value"],
        ["Amerin Gauge Systems", "Engineered measurement components", "$105,000", "$128,000"],
        ["Caldera Flow Controls", "Specialty valves and flow control", "$142,000", "$181,000"],
        ["Northline Fabrication", "Precision metal fabrication", "$84,000", "$92,000"],
        ["Vantage Field Services", "Facility compliance services", "$119,000", "$155,000"],
        ["Other investments", "Various", "$75,000", "$54,000"],
        ["Total", "", "$525,000", "$610,000"],
    ], widths=[1.8 * inch, 1.8 * inch, 1.2 * inch, 1.2 * inch]))
    maybe_page_break(story)
    section("5", "NAV Roll-Forward", [
        "Capital activity and portfolio appreciation during 2025 supported Fund III's trajectory into the 2026 reporting period. Cumulative paid-in capital at year-end was approximately $525.0 million, cumulative distributions were $170.0 million and NAV was $610.0 million.",
        *boilerplate("NAV roll-forward", 2),
    ], story)
    story.append(table([
        ["Roll-forward", "Amount"],
        ["Beginning partners' capital", "$420,400"],
        ["Capital contributions", "$90,000"],
        ["Distributions", "($55,000)"],
        ["Management fees and expenses", "($21,100)"],
        ["Realized and unrealized gains", "$175,700"],
        ["Ending partners' capital / NAV", "$610,000"],
    ], widths=[3.0 * inch, 2.0 * inch]))
    maybe_page_break(story)
    section("6", "Notes to Financial Statements", [
        "Investments are carried at fair value as determined in good faith by the General Partner under the Partnership's valuation policy. The Partnership uses Level 3 inputs for privately held portfolio companies.",
        "Management fees are charged in accordance with the Partnership Agreement. Transaction and monitoring fee offsets are applied based on governing documents and related disclosures.",
        "The Partnership may use a subscription facility to bridge capital calls and manage short-term liquidity. Borrowings are expected to be repaid from limited partner capital contributions.",
        *boilerplate("Notes to financial statements", 3),
    ], story)
    build_pdf("brightwater_iii_audited_fs_2025.pdf", "Brightwater Capital Partners III, L.P.", "Audited Financial Statements - FY2025 | Vyntic token: financial_statements", story)


def generate_manifest(run_date: dt.date):
    due = run_date + dt.timedelta(days=5)
    pay = run_date + dt.timedelta(days=12)
    manifest = f"""# Brightwater / Glenmoor Demo Document Manifest

{SAMPLE_LINE}

All documents in this folder are realistic, fictional private-equity fund documents generated for Vyntic demo and workflow testing. No real firm, fund, portfolio company, person, bank, auditor, law firm or investor is represented.

## Install Commands Run

No library installation commands were run. The local environment already had `reportlab` and `openpyxl` available.

If reproducing in a clean Python environment, install the required libraries with:

```bash
python3 -m pip install reportlab openpyxl
```

Generation command used:

```bash
python3 scripts/generate_brightwater_docs.py --run-date {run_date.isoformat()}
```

## Consistency Ledger

- GP / Manager: Brightwater Capital Partners, LLC; Chicago; founded 2009; approximately $2.1B AUM; North American industrials and business services.
- Our LP: Glenmoor University Endowment.
- Fund III: Brightwater Capital Partners III, L.P.; 2021 vintage; $850M fund size; European waterfall; 8% preferred return; 20% carry.
- Fund IV: Brightwater Capital Partners IV, L.P.; 2026 vintage; $1.25B target; $1.5B hard cap.
- Glenmoor Fund III Q2 2026 position: $25,000,000 commitment; $18,750,000 paid-in / called; $6,200,000 distributed; $21,400,000 NAV; DPI 0.33x; RVPI 1.14x; TVPI 1.47x.
- Capital Call No. 7: issue date {date_text(run_date)}; due date {date_text(due)}; amount $1,875,000; remaining unfunded commitment after call $4,375,000.
- Distribution Notice No. 3: notice date {date_text(run_date)}; pay date {date_text(pay)}; amount $1,400,000.
- Q2 2026 quarterly report: quarter-end June 30, 2026; report date August 29, 2026.

## Track Record Ledger

| Fund | Vintage | Fund Size | Net IRR | TVPI | DPI | RVPI | Expected Tie |
| --- | ---: | ---: | ---: | ---: | ---: | ---: | --- |
| Fund I | 2011 | $300,000,000 | 18% | 2.10x | 2.10x | 0.00x | Ties |
| Fund II | 2015 | $550,000,000 | 21% | 1.90x | 1.30x | 0.60x | Ties |
| Fund III | 2021 | $850,000,000 | 15% | 1.50x | 0.40x | 0.95x | Planted mismatch: 0.40x + 0.95x = 1.35x, not 1.50x |

## Files and Expected Findings

| Filename | Vyntic token | Workflow / demo exercised | Expected findings |
| --- | --- | --- | --- |
| brightwater_iv_ppm.pdf | ppm | Fund IV selection diligence; terms, team and performance extraction | Lists Daniel Roache as active senior team; includes 50% fee offset, 2.0% GP commitment, 80% no-fault removal threshold, $2.5M organizational expense cap; prior fund summary presents Fund III headline at 1.50x. |
| brightwater_iv_pitchbook.pdf | pitchbook | Marketing claim and track-record validation | Claims top-quartile track record and 2.1x net using fully realized Fund I; lists Daniel Roache as active; shows Fund III at 1.50x. |
| brightwater_track_record.xlsx | track_record | Spreadsheet extraction and cross-checking | Fund III planted mismatch: TVPI 1.50x while DPI 0.40x plus RVPI 0.95x equals 1.35x; Fund I and Fund II ties are correct. |
| brightwater_iv_ddq.pdf | ddq | DDQ inconsistency and ODD workflow | Team / succession answer is evasive and does not disclose Daniel Roache's departure; fees answer says 100% fee offset but LPA provides only 50%; conflicts answer omits Brightwater Securities, LLC; cybersecurity answer lacks SOC 2 and fixed pen-test cadence. |
| brightwater_iv_lpa.pdf | lpa | Fund terms extraction and off-market term detection | Fee offset only 50%; Daniel Roache named as a Key Person; no-fault GP removal requires 80%; GP commitment only 2.0%; recycling for 24 months; organizational expense cap $2,500,000. |
| brightwater_adv_part2a.pdf | form_adv | Regulatory and conflict cross-checking | Discloses affiliated broker-dealer Brightwater Securities, LLC receiving transaction fees; discloses 2023 SEC deficiency letter on expense allocation, remediated; discloses Daniel Roache ceased being an advisory employee effective February 28, 2026. |
| brightwater_valuation_policy.pdf | valuation_policy | ODD valuation governance workflow | Quarterly valuation committee, but Level 3 assets are GP-marked with third-party review only annually. |
| glenmoor_fund_iii_side_letter.pdf | side_letter | Side-letter obligation extraction | Obligations: 10 bps fee reduction; MFN right; pro-rata co-investment offer above $75M equity; quarterly reports within 45 days and audited annuals within 120 days; tobacco / controlled-substance excuse; affiliate transfer consent not unreasonably withheld; annual ESG report. |
| glenmoor_fund_iii_pcap_q2_2026.pdf | capital_account | Monitoring position and fee-compliance workflow | Exact Glenmoor position ties to ledger; management fee allocation reflects 10 bps discount, so side-letter obligation (i) is compliant. |
| brightwater_iii_quarterly_q2_2026.pdf | quarterly_report | Monitoring package extraction and side-letter testing | Report dated August 29, 2026, 60 days after June 30, 2026, breaching 45-day quarterly reporting undertaking; omits portfolio-level ESG metrics, so annual ESG-report obligation is unclear / potential breach; Glenmoor position ties to PCAP and ledger. |
| brightwater_iii_capital_call_07.pdf | capital_call | Capital call extraction and urgency detection | Amount $1,875,000; due {date_text(due)}, five calendar days after issue; purpose Project Cardinal; remaining unfunded commitment after call $4,375,000. |
| brightwater_iii_distribution_03.pdf | distribution_notice | Distribution notice extraction | Amount $1,400,000 split $950,000 return of capital and $450,000 gain; pay date {date_text(pay)}. |
| brightwater_iii_audited_fs_2025.pdf | financial_statements | Audited financial statement extraction and NAV trajectory | Fictional auditor opinion; FY2025 schedule and NAV roll-forward support Fund III trajectory into Q2 2026; Level 3 valuation note. |

## Side-Letter Verdicts Against Q2 2026 Monitoring Pack

| Side-letter obligation | Expected verdict | Reason |
| --- | --- | --- |
| (i) 10 bps management-fee reduction | Compliant | PCAP and quarterly report show Glenmoor fee allocation at 1.90%, equal to 2.00% less 10 bps. |
| (ii) MFN election right | Unclear | Not directly verifiable from Q2 report or notices. |
| (iii) Pro-rata co-investment offer over $75M equity | Unclear | Project Cardinal call does not provide enough information about aggregate equity or co-investment process. |
| (iv) Quarterly reports within 45 days | Breach | Q2 report is dated August 29, 2026, 60 days after June 30, 2026. |
| (v) Tobacco / controlled-substance excuse right | Unclear | No covered investment appears in the Q2 report. |
| (vi) Affiliate transfer consent not unreasonably withheld | Unclear | No transfer request is shown. |
| (vii) Annual ESG report | Unclear / potential breach | Q2 report contains no portfolio-level ESG metrics; annual compliance cannot be fully verified from Q2 alone. |
"""
    (OUTPUT / "MANIFEST.md").write_text(manifest, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-date", default="2026-07-22", help="Issue/notice date in YYYY-MM-DD format")
    args = parser.parse_args()
    run_date = dt.date.fromisoformat(args.run_date)
    OUTPUT.mkdir(parents=True, exist_ok=True)

    generate_ppm(run_date)
    generate_pitchbook()
    generate_track_record_xlsx()
    generate_ddq(run_date)
    generate_lpa()
    generate_adv(run_date)
    generate_valuation_policy()
    generate_side_letter()
    generate_pcap()
    generate_quarterly()
    generate_capital_call(run_date)
    generate_distribution(run_date)
    generate_financial_statements()
    generate_manifest(run_date)

    print(f"Generated Brightwater demo pack in {OUTPUT}")


if __name__ == "__main__":
    main()
