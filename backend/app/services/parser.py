"""
Document parser: converts PDFs and Excel files into structured Markdown sections.
Uses docling for PDFs (high-quality table + text extraction) and openpyxl for Excel.
"""
import logging
import os
import uuid
import gc
import json
import asyncio
import multiprocessing
import tempfile
import traceback
import threading
from pathlib import Path
from collections.abc import Callable

_log = logging.getLogger(__name__)

from app.config import settings
from app.models.document import ParsedSection, DocumentMetadata

_docling_job_semaphore = threading.Semaphore(max(1, settings.docling_max_concurrent_jobs))


def _configure_docling_runtime() -> None:
    """Keep Docling's model stack conservative on local/macOS startup."""
    os.environ.setdefault("OMP_NUM_THREADS", str(settings.docling_num_threads))
    os.environ.setdefault("MKL_NUM_THREADS", str(settings.docling_num_threads))
    os.environ.setdefault("NUMEXPR_NUM_THREADS", str(settings.docling_num_threads))
    os.environ.setdefault("VECLIB_MAXIMUM_THREADS", str(settings.docling_num_threads))
    os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")


def _set_docling_option(options, name: str, value) -> None:
    """Set a Docling option only when the installed version exposes it."""
    if hasattr(options, name):
        setattr(options, name, value)


def _docling_convert_pdf(
    file_path: str,
    page_range: tuple[int, int] | None = None,
) -> list[dict]:
    """Run Docling and return plain page data that can cross process boundaries."""
    _configure_docling_runtime()

    from docling.document_converter import DocumentConverter
    from docling.document_converter import PdfFormatOption
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import (
        AcceleratorDevice,
        AcceleratorOptions,
        PdfPipelineOptions,
    )
    from docling_core.types.doc import TableItem, TextItem

    device_name = settings.docling_device.lower()
    device = AcceleratorDevice.CPU
    if device_name in {item.value for item in AcceleratorDevice}:
        device = AcceleratorDevice(device_name)

    pipeline_options = PdfPipelineOptions(
        accelerator_options=AcceleratorOptions(
            num_threads=settings.docling_num_threads,
            device=device,
        ),
        document_timeout=settings.docling_timeout_seconds,
        do_ocr=settings.docling_ocr_enabled,
    )
    for option_name, option_value in {
        "ocr_batch_size": 1,
        "layout_batch_size": 1,
        "table_batch_size": 1,
        "queue_max_size": settings.docling_queue_max_size,
        "generate_page_images": False,
        "generate_picture_images": False,
        "generate_table_images": False,
        "generate_parsed_pages": False,
        "images_scale": 0.5,
        "force_backend_text": True,
    }.items():
        _set_docling_option(pipeline_options, option_name, option_value)

    converter = DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)
        }
    )
    convert_kwargs = {}
    if page_range is not None:
        convert_kwargs["page_range"] = page_range
    result = converter.convert(file_path, **convert_kwargs)
    doc = result.document

    pages: dict[int, dict] = {}
    for item, _level in doc.iterate_items():
        if hasattr(item, "prov") and item.prov:
            page_num = item.prov[0].page_no
        else:
            page_num = 1

        if page_num not in pages:
            pages[page_num] = {"text": [], "tables": [], "has_table": False}

        if isinstance(item, TableItem):
            md_table = item.export_to_markdown(doc=doc)
            if md_table and md_table.strip():
                pages[page_num]["tables"].append(md_table)
                pages[page_num]["has_table"] = True
        elif isinstance(item, TextItem):
            text = item.text
            if text and text.strip():
                pages[page_num]["text"].append(text.strip())

    return [
        {
            "page_number": page_num,
            "text": pages[page_num]["text"],
            "tables": pages[page_num]["tables"],
            "has_table": pages[page_num]["has_table"],
        }
        for page_num in sorted(pages.keys())
    ]


def _docling_worker(
    file_path: str,
    output_path: str,
    page_start: int | None = None,
    page_end: int | None = None,
) -> None:
    try:
        page_range = None
        if page_start is not None and page_end is not None:
            page_range = (page_start, page_end)
        payload = {
            "status": "ok",
            "pages": _docling_convert_pdf(file_path, page_range=page_range),
        }
    except Exception:
        payload = {"status": "error", "error": traceback.format_exc()}
    Path(output_path).write_text(json.dumps(payload), encoding="utf-8")
    gc.collect()


def _read_docling_worker_output(output_path: Path) -> list[dict]:
    if not output_path.exists() or output_path.stat().st_size == 0:
        raise FileNotFoundError("Docling worker produced no parse output")

    payload = json.loads(output_path.read_text(encoding="utf-8"))
    if payload["status"] == "error":
        raise RuntimeError(payload["error"])
    return payload["pages"]


def _docling_process_context() -> multiprocessing.context.BaseContext:
    return multiprocessing.get_context("spawn")


def _run_docling_worker(
    file_path: Path,
    page_range: tuple[int, int] | None = None,
) -> list[dict]:
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as tmp:
        output_path = Path(tmp.name)

    try:
        page_start = page_range[0] if page_range else None
        page_end = page_range[1] if page_range else None
        ctx = _docling_process_context()
        proc = ctx.Process(
            target=_docling_worker,
            args=(str(file_path), str(output_path), page_start, page_end),
        )
        proc.start()
        proc.join(
            settings.docling_timeout_seconds + settings.docling_worker_grace_seconds
        )

        if proc.is_alive():
            proc.terminate()
            proc.join(10)
            if proc.is_alive():
                proc.kill()
                proc.join(10)
            raise TimeoutError(
                f"Docling timed out after {settings.docling_timeout_seconds}s"
            )

        if proc.exitcode != 0:
            raise RuntimeError(
                f"Docling worker exited unexpectedly with code {proc.exitcode}"
            )

        return _read_docling_worker_output(output_path)
    finally:
        try:
            output_path.unlink()
        except FileNotFoundError:
            pass


def _count_pdf_pages(file_path: Path) -> int | None:
    try:
        import pypdfium2 as pdfium

        pdf = pdfium.PdfDocument(str(file_path))
        try:
            return len(pdf)
        finally:
            close = getattr(pdf, "close", None)
            if close:
                close()
    except Exception:
        return None


def _convert_pdf_range_with_retries(
    file_path: Path,
    start_page: int,
    end_page: int,
) -> list[dict]:
    try:
        return _run_docling_worker(file_path, page_range=(start_page, end_page))
    except Exception:
        if start_page >= end_page:
            raise

        midpoint = (start_page + end_page) // 2
        return (
            _convert_pdf_range_with_retries(file_path, start_page, midpoint)
            + _convert_pdf_range_with_retries(file_path, midpoint + 1, end_page)
        )


def _dedupe_pages(pages: list[dict]) -> list[dict]:
    deduped: dict[int, dict] = {}
    for page in pages:
        page_number = int(page["page_number"])
        if page_number not in deduped:
            deduped[page_number] = page
            continue
        deduped[page_number]["text"].extend(page.get("text", []))
        deduped[page_number]["tables"].extend(page.get("tables", []))
        deduped[page_number]["has_table"] = (
            deduped[page_number]["has_table"] or page.get("has_table", False)
        )
    return [deduped[page_num] for page_num in sorted(deduped)]


def _pages_to_full_text_md(pages: list[dict]) -> str:
    """Convert raw pages dict list to the '## Page N' markdown format stored in full_text_md."""
    parts = []
    for page in pages:
        content_parts = []
        for text in page.get("text", []):
            if text and text.strip():
                content_parts.append(text.strip())
        for table in page.get("tables", []):
            if table and table.strip():
                content_parts.append(table.strip())
        if not content_parts:
            continue
        parts.append(f"## Page {page['page_number']}")
        parts.extend(content_parts)
    return "\n\n".join(parts)


def _pymupdf_parse_pdf(file_path: Path) -> list[dict]:
    """Parse PDF using PyMuPDF. Native text extraction only — no OCR.

    Returns pages in the same format as Docling (list of page dicts).
    Fast (~1s / 100 pages). Used as Tier 2 fallback.
    """
    import pymupdf

    pages = []
    doc = pymupdf.open(str(file_path))
    try:
        for page_index in range(len(doc)):
            page = doc[page_index]
            page_number = page_index + 1
            text = page.get_text("text")
            if text and text.strip():
                pages.append({
                    "page_number": page_number,
                    "text": [text.strip()],
                    "tables": [],
                    "has_table": False,
                })
    finally:
        doc.close()

    return pages


_FULL_TEXT_MIN_CHARS = 100


def _parse_with_cascade(
    file_path: Path,
    progress_callback: Callable[[float, str], None] | None = None,
) -> tuple[list[dict], int]:
    """Run the 3-tier parsing cascade. Returns (pages, parse_tier).

    Tier 1: Docling (default) — succeeds if no exception AND total chars >= 100.
    Tier 2: PyMuPDF — fast native text extraction, no OCR.
    Tier 3: Azure Document Intelligence — credential-gated; only attempted if
            AZURE_DI_ENDPOINT and AZURE_DI_KEY are set.

    parse_tier value: 1=Docling, 2=PyMuPDF, 3=AzureDI.
    """
    # Tier 1: Docling
    try:
        pages = _convert_pdf_isolated_with_progress(file_path, progress_callback)
        total_chars = sum(len(t) for p in pages for t in p.get("text", []) + p.get("tables", []))
        if total_chars >= _FULL_TEXT_MIN_CHARS:
            return pages, 1
        _log.warning(
            "Docling produced only %d chars for %s — falling back to PyMuPDF",
            total_chars,
            file_path.name,
        )
    except Exception as e:
        _log.warning("Docling failed for %s: %s — falling back to PyMuPDF", file_path.name, e)

    # Tier 2: PyMuPDF
    try:
        pages = _pymupdf_parse_pdf(file_path)
        pymupdf_chars = sum(len(t) for p in pages for t in p.get("text", []) + p.get("tables", []))
        if pymupdf_chars >= _FULL_TEXT_MIN_CHARS:
            return pages, 2
        _log.warning(
            "PyMuPDF produced only %d chars for %s — falling back to Azure DI",
            pymupdf_chars,
            file_path.name,
        )
        pymupdf_exc = ValueError(f"PyMuPDF output too short ({pymupdf_chars} chars) for {file_path.name}")
    except Exception as e:
        _log.warning("PyMuPDF failed for %s: %s — falling back to Azure DI", file_path.name, e)
        pymupdf_exc = e

    # Tier 3: Azure Document Intelligence (credential-gated)
    import os
    if not (os.environ.get("AZURE_DI_ENDPOINT") and os.environ.get("AZURE_DI_KEY")):
        raise ValueError(
            f"PDF parsing failed for {file_path.name}: Docling and PyMuPDF both failed, "
            "and Azure DI credentials (AZURE_DI_ENDPOINT, AZURE_DI_KEY) are not configured."
        ) from pymupdf_exc

    try:
        pages = _azure_di_parse_pdf(file_path)
        return pages, 3
    except Exception as e:
        raise ValueError(
            f"All three parsing tiers failed for {file_path.name}. Last error: {e}"
        ) from e


def _azure_di_parse_pdf(file_path: Path) -> list[dict]:
    """Parse PDF using Azure Document Intelligence prebuilt-layout model.

    Requires AZURE_DI_ENDPOINT and AZURE_DI_KEY env vars.
    Returns pages in the same format as Docling and PyMuPDF tiers.
    """
    import os
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential

    endpoint = os.environ["AZURE_DI_ENDPOINT"]
    key = os.environ["AZURE_DI_KEY"]

    with DocumentAnalysisClient(endpoint, AzureKeyCredential(key)) as client:
        with open(str(file_path), "rb") as f:
            poller = client.begin_analyze_document("prebuilt-layout", document=f)
        result = poller.result()

    pages: dict[int, dict] = {}

    for page in result.pages:
        pn = page.page_number
        pages[pn] = {"page_number": pn, "text": [], "tables": [], "has_table": False}
        for line in (page.lines or []):
            if line.content and line.content.strip():
                pages[pn]["text"].append(line.content.strip())

    for table in (result.tables or []):
        if not table.bounding_regions:
            _log.warning("Azure DI table has no bounding_regions — skipping")
            continue
        pn = table.bounding_regions[0].page_number
        if pn not in pages:
            pages[pn] = {"page_number": pn, "text": [], "tables": [], "has_table": False}
        num_cols = table.column_count
        num_rows = table.row_count
        cells = [[""] * num_cols for _ in range(num_rows)]
        for cell in table.cells:
            if 0 <= cell.row_index < num_rows and 0 <= cell.column_index < num_cols:
                cells[cell.row_index][cell.column_index] = cell.content or ""
        if cells:
            md_rows = ["| " + " | ".join(cells[0]) + " |"]
            md_rows.append("| " + " | ".join(["---"] * num_cols) + " |")
            for row in cells[1:]:
                md_rows.append("| " + " | ".join(row) + " |")
            pages[pn]["tables"].append("\n".join(md_rows))
            pages[pn]["has_table"] = True

    return [pages[n] for n in sorted(pages.keys())]


def _convert_pdf_isolated(file_path: Path) -> list[dict]:
    page_count = _count_pdf_pages(file_path)
    batch_size = max(1, settings.docling_page_batch_size)

    if not page_count or page_count <= batch_size:
        return _run_docling_worker(file_path)

    pages: list[dict] = []
    for start_page in range(1, page_count + 1, batch_size):
        end_page = min(start_page + batch_size - 1, page_count)
        pages.extend(
            _convert_pdf_range_with_retries(file_path, start_page, end_page)
        )

    return _dedupe_pages(pages)


def _convert_pdf_isolated_with_progress(
    file_path: Path,
    progress_callback: Callable[[float, str], None] | None = None,
) -> list[dict]:
    page_count = _count_pdf_pages(file_path)
    batch_size = max(1, settings.docling_page_batch_size)

    if not page_count or page_count <= batch_size:
        pages = _run_docling_worker(file_path)
        if progress_callback:
            progress_callback(1.0, "Parsed document")
        return pages

    pages: list[dict] = []
    total_batches = (page_count + batch_size - 1) // batch_size
    for batch_index, start_page in enumerate(range(1, page_count + 1, batch_size), 1):
        end_page = min(start_page + batch_size - 1, page_count)
        pages.extend(
            _convert_pdf_range_with_retries(file_path, start_page, end_page)
        )
        if progress_callback:
            progress_callback(
                batch_index / total_batches,
                f"Parsed pages {start_page}-{end_page} of {page_count}",
            )

    return _dedupe_pages(pages)


def _convert_pdf_isolated_with_lock(
    file_path: Path,
    progress_callback: Callable[[float, str], None] | None = None,
) -> list[dict]:
    with _docling_job_semaphore:
        return _convert_pdf_isolated_with_progress(file_path, progress_callback)


def _docling_convert_pdf_with_lock(file_path: str) -> list[dict]:
    with _docling_job_semaphore:
        return _docling_convert_pdf(file_path)


def _build_pdf_sections(
    doc_id: str,
    filename: str,
    deal_id: str,
    pages: list[dict],
) -> list[ParsedSection]:
    # Emit prose and each table as separate sections so a citation can point
    # at the specific table the LLM used, instead of a mixed page blob whose
    # snippet would only show the page preamble.
    sections = []
    for page_data in pages:
        text_content = "\n\n".join(page_data["text"]).strip()
        if text_content:
            sections.append(
                ParsedSection(
                    content=text_content,
                    metadata={
                        "source_file": filename,
                        "page_number": page_data["page_number"],
                        "section_type": "text",
                        "deal_id": deal_id,
                        "doc_id": doc_id,
                    },
                )
            )
        for table_md in page_data["tables"]:
            table_content = table_md.strip()
            if not table_content:
                continue
            sections.append(
                ParsedSection(
                    content=table_content,
                    metadata={
                        "source_file": filename,
                        "page_number": page_data["page_number"],
                        "section_type": "table",
                        "deal_id": deal_id,
                        "doc_id": doc_id,
                    },
                )
            )
    return sections


async def parse_pdf_path(
    file_path: Path,
    filename: str,
    deal_id: str,
    progress_callback: Callable[[float, str], None] | None = None,
) -> tuple[DocumentMetadata, list[ParsedSection]]:
    """Parse a PDF file using the 3-tier cascade (Docling → PyMuPDF → Azure DI).

    Sets full_text_md and parse_tier on the returned DocumentMetadata.
    """
    doc_id = f"{deal_id}_{uuid.uuid4().hex[:8]}"

    pages, parse_tier = await asyncio.to_thread(
        _parse_with_cascade, file_path, progress_callback
    )

    sections = _build_pdf_sections(doc_id, filename, deal_id, pages)
    full_text_md = _pages_to_full_text_md(pages)
    detected_page_count = _count_pdf_pages(file_path)
    parsed_page_count = max((int(p["page_number"]) for p in pages), default=0)

    metadata = DocumentMetadata(
        doc_id=doc_id,
        deal_id=deal_id,
        filename=filename,
        page_count=detected_page_count or parsed_page_count,
        full_text_md=full_text_md or None,
        parse_tier=parse_tier,
    )

    return metadata, sections


async def parse_pdf(
    file_bytes: bytes,
    filename: str,
    deal_id: str,
) -> tuple[DocumentMetadata, list[ParsedSection]]:
    """Parse PDF bytes using Docling, extracting text and tables as Markdown."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        return await parse_pdf_path(Path(tmp_path), filename, deal_id)
    finally:
        os.unlink(tmp_path)


async def parse_excel(
    file_bytes: bytes,
    filename: str,
    deal_id: str,
) -> tuple[DocumentMetadata, list[ParsedSection]]:
    """Parse an Excel file using openpyxl, converting each sheet to Markdown tables."""
    import openpyxl
    from io import BytesIO

    doc_id = f"{deal_id}_{uuid.uuid4().hex[:8]}"
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Build Markdown table
        md_lines = [f"## {sheet_name}\n"]

        # Header row
        headers = [str(c) if c is not None else "" for c in rows[0]]
        md_lines.append("| " + " | ".join(headers) + " |")
        md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        # Data rows
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            while len(cells) < len(headers):
                cells.append("")
            md_lines.append("| " + " | ".join(cells[:len(headers)]) + " |")

        content = "\n".join(md_lines)
        sections.append(
            ParsedSection(
                content=content,
                metadata={
                    "source_file": filename,
                    "page_number": 1,
                    "section_type": "table",
                    "sheet_name": sheet_name,
                    "deal_id": deal_id,
                    "doc_id": doc_id,
                },
            )
        )

    metadata = DocumentMetadata(
        doc_id=doc_id,
        deal_id=deal_id,
        filename=filename,
        page_count=len(sections),
    )

    return metadata, sections


async def parse_document(
    file_bytes: bytes,
    filename: str,
    deal_id: str,
) -> tuple[DocumentMetadata, list[ParsedSection]]:
    """Route to the appropriate parser based on file extension."""
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return await parse_pdf(file_bytes, filename, deal_id)
    elif ext in (".xlsx", ".xls"):
        return await parse_excel(file_bytes, filename, deal_id)
    else:
        raise ValueError(
            f"Unsupported file type: {ext}. Supported: .pdf, .xlsx, .xls"
        )


async def parse_document_path(
    file_path: Path,
    filename: str,
    deal_id: str,
    progress_callback: Callable[[float, str], None] | None = None,
) -> tuple[DocumentMetadata, list[ParsedSection]]:
    """Route to the appropriate parser for a file already on disk."""
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return await parse_pdf_path(file_path, filename, deal_id, progress_callback)
    elif ext in (".xlsx", ".xls"):
        return await parse_excel(file_path.read_bytes(), filename, deal_id)
    else:
        raise ValueError(
            f"Unsupported file type: {ext}. Supported: .pdf, .xlsx, .xls"
        )
