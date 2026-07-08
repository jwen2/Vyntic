"""Deal CRUD routes."""
import os
import shutil
from html import escape
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import FileResponse

from app.config import settings
from app.models.deal import (
    Deal,
    DealCreate,
    DealUpdate,
    DEAL_STAGES,
    FUND_STAGES,
    ENTITY_TYPES,
    DOC_CATEGORIES,
    DOC_SCOPES,
    SECTOR_TAGS,
    stages_for_entity,
)
from app.models.document import DocumentMetadata
from app.models.manager import Position, PositionUpsert
from app.services import audit_store, deal_store, manager_store
from app.database import UserRow
from app.auth import (
    create_scoped_token,
    doc_view_query_auth,
    get_current_user,
    grant_deal_access,
    require_admin,
    require_deal_access,
)

router = APIRouter(prefix="/deals", tags=["deals"])

# Mounted WITHOUT the app-wide get_current_user dependency (main.py): the
# document viewer authenticates via ?token= for iframes, where the browser
# cannot set an Authorization header. Every route here must carry
# get_current_user_or_query_token explicitly; the default-deny walker test
# (tests/test_default_deny.py) enforces that nothing on it is open.
view_router = APIRouter(prefix="/deals", tags=["deals"])


@router.post("", response_model=Deal)
def create_deal(
    data: DealCreate,
    http_request: Request,
    current_user: UserRow = Depends(get_current_user),
):
    require_admin(current_user)
    if data.entity_type not in ENTITY_TYPES:
        raise HTTPException(status_code=422, detail=f"entity_type must be one of {ENTITY_TYPES}")
    if data.stage not in stages_for_entity(data.entity_type):
        raise HTTPException(
            status_code=422,
            detail=f"Invalid stage '{data.stage}' for entity_type '{data.entity_type}'",
        )
    if data.manager_id:
        if data.entity_type != "fund":
            raise HTTPException(status_code=422, detail="Only funds can belong to a manager")
        # Tenant-scoped lookup: another tenant's manager is invisible here.
        if not manager_store.get_manager(data.manager_id, tenant_id=current_user.tenant_id):
            raise HTTPException(status_code=422, detail=f"Manager '{data.manager_id}' not found")
    try:
        deal = deal_store.create_deal(data, tenant_id=current_user.tenant_id)
        # Auto-grant access to the creator
        grant_deal_access(current_user.id, data.deal_id, role="admin")
        audit_store.record(
            current_user, "deal.create", resource_type="deal",
            resource_id=data.deal_id, deal_id=data.deal_id, request=http_request,
        )
        return deal
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))


@router.get("", response_model=list[Deal])
def list_deals(current_user: UserRow = Depends(get_current_user)):
    return deal_store.list_deals(tenant_id=current_user.tenant_id)


@router.get("/metadata/stages")
def get_stages(entity_type: str = "deal"):
    """Return valid pipeline stages. Defaults to the buyout-deal pipeline for
    backwards compatibility; pass ?entity_type=fund for the LP fund lifecycle."""
    if entity_type not in ENTITY_TYPES:
        raise HTTPException(status_code=422, detail=f"entity_type must be one of {ENTITY_TYPES}")
    return stages_for_entity(entity_type)


@router.get("/metadata/tags")
def get_tags():
    """Return suggested sector tags."""
    return SECTOR_TAGS


@router.get("/metadata/doc-categories")
def get_doc_categories():
    """Return the LP document taxonomy for classification dropdowns."""
    return DOC_CATEGORIES


@router.get("/{deal_id}", response_model=Deal)
def get_deal(deal_id: str, current_user: UserRow = Depends(get_current_user)):
    require_deal_access(current_user, deal_id)
    deal = deal_store.get_deal(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")
    return deal


@router.patch("/{deal_id}", response_model=Deal)
def update_deal(deal_id: str, data: DealUpdate, current_user: UserRow = Depends(get_current_user)):
    require_deal_access(current_user, deal_id)
    existing = deal_store.get_deal(deal_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")
    if data.stage is not None:
        # Stage moves are admin-only per the access model; analysts may still
        # edit name/description/tags on deals they can access.
        require_admin(current_user)
        if data.stage not in stages_for_entity(existing.entity_type):
            raise HTTPException(
                status_code=422,
                detail=f"Invalid stage '{data.stage}' for entity_type '{existing.entity_type}'",
            )
    if data.manager_id is not None:
        require_admin(current_user)
        if existing.entity_type != "fund":
            raise HTTPException(status_code=422, detail="Only funds can belong to a manager")
        if not manager_store.get_manager(data.manager_id):
            raise HTTPException(status_code=422, detail=f"Manager '{data.manager_id}' not found")
    deal = deal_store.update_deal(deal_id, data)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")
    return deal


# ── Position (the LP's commitment in a fund) ──

@router.get("/{deal_id}/position", response_model=Position)
def get_position(deal_id: str, current_user: UserRow = Depends(get_current_user)):
    require_deal_access(current_user, deal_id)
    if not deal_store.get_deal(deal_id):
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")
    position = manager_store.get_position(deal_id)
    if not position:
        # Empty position rather than 404 — the UI treats "no position yet"
        # as an editable blank form, not an error.
        return Position(deal_id=deal_id)
    return position


@router.put("/{deal_id}/position", response_model=Position)
def upsert_position(
    deal_id: str,
    data: PositionUpsert,
    current_user: UserRow = Depends(get_current_user),
):
    require_admin(current_user)
    require_deal_access(current_user, deal_id)
    deal = deal_store.get_deal(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")
    if deal.entity_type != "fund":
        raise HTTPException(status_code=422, detail="Positions only apply to funds")
    return manager_store.upsert_position(deal_id, data)


@router.get("/{deal_id}/documents", response_model=list[DocumentMetadata])
def list_deal_documents(deal_id: str, current_user: UserRow = Depends(get_current_user)):
    require_deal_access(current_user, deal_id)
    deal = deal_store.get_deal(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")
    return deal_store.list_documents(deal_id)


@router.delete("/{deal_id}")
async def delete_deal(
    deal_id: str,
    http_request: Request,
    current_user: UserRow = Depends(get_current_user),
):
    require_admin(current_user)
    require_deal_access(current_user, deal_id)
    from app.services.vector_store import delete_deal_vectors

    if not deal_store.delete_deal(deal_id):
        raise HTTPException(status_code=404, detail=f"Deal '{deal_id}' not found")

    audit_store.record(
        current_user, "deal.delete", resource_type="deal",
        resource_id=deal_id, deal_id=deal_id, request=http_request,
    )

    try:
        await delete_deal_vectors(deal_id)
    except Exception:
        pass  # Best-effort cleanup of vectors

    # Clean up uploaded files
    deal_upload_dir = os.path.join(settings.uploads_dir, deal_id)
    if os.path.isdir(deal_upload_dir):
        shutil.rmtree(deal_upload_dir)

    return {"status": "deleted", "deal_id": deal_id}


@router.get("/{deal_id}/documents/{filename}/view-token")
def mint_view_token(
    deal_id: str,
    filename: str,
    current_user: UserRow = Depends(get_current_user),
):
    """Mint a short-lived token scoped to viewing exactly this document.
    The viewer iframe carries it in ?token= (browsers can't set headers
    there); it cannot be replayed for other files or as a session token."""
    require_deal_access(current_user, deal_id)
    token = create_scoped_token(
        "doc-view", {"deal_id": deal_id, "filename": filename}, user_id=current_user.id
    )
    return {"token": token, "expires_in": 300}


@view_router.get("/{deal_id}/documents/{filename}/view")
async def view_document(
    deal_id: str,
    filename: str,
    http_request: Request,
    sheet: int | None = None,
    token: str | None = None,
    current_user: UserRow = Depends(doc_view_query_auth),
):
    require_deal_access(current_user, deal_id)
    audit_store.record(
        current_user, "document.view", resource_type="document",
        resource_id=filename, deal_id=deal_id, request=http_request,
    )
    """Serve an original uploaded document file for inline viewing (not download).

    For Excel files, optionally pass ?sheet=0 to view a specific sheet.
    Large sheets are truncated to the first 500 rows.
    """
    file_path = os.path.join(settings.uploads_dir, deal_id, filename)
    if not os.path.exists(file_path):
        # Manager-scoped documents live in the sibling fund they were uploaded
        # to but appear in this fund's context — resolve the shared file so
        # citation clicks keep working. Access to this fund implies access to
        # the manager's shared documents (same rule as context assembly).
        shared_path = _resolve_manager_shared_file(deal_id, filename)
        if shared_path is None:
            raise HTTPException(status_code=404, detail="Document file not found")
        file_path = shared_path

    lower = filename.lower()

    # Excel files: convert to HTML table for inline preview
    if lower.endswith((".xlsx", ".xls")):
        return _excel_to_html_response(file_path, filename, active_sheet=sheet, token=token)

    if lower.endswith(".pdf"):
        media_type = "application/pdf"
    elif lower.endswith(".docx"):
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif lower.endswith(".txt"):
        media_type = "text/plain"
    elif lower.endswith(".csv"):
        media_type = "text/csv"
    else:
        # Unknown/HTML-ish types must never render in the app origin — a
        # renamed .html would script against the app. Force download.
        return FileResponse(
            file_path,
            media_type="application/octet-stream",
            filename=filename,
            content_disposition_type="attachment",
            headers={"X-Content-Type-Options": "nosniff"},
        )

    # content_disposition_type="inline" prevents download — renders in browser
    return FileResponse(
        file_path,
        media_type=media_type,
        filename=filename,
        content_disposition_type="inline",
        headers={"X-Content-Type-Options": "nosniff"},
    )


def _resolve_manager_shared_file(deal_id: str, filename: str) -> str | None:
    """Locate a manager-scoped document's file via the sibling fund it was
    uploaded to. Returns None when the deal has no manager or no sibling holds
    a manager-scoped document with this filename."""
    deal = deal_store.get_deal(deal_id)
    if not deal or not deal.manager_id:
        return None
    for doc in deal_store.list_manager_documents(deal.manager_id):
        if doc.filename == filename:
            candidate = os.path.join(settings.uploads_dir, doc.deal_id, filename)
            if os.path.exists(candidate):
                return candidate
    return None


MAX_PREVIEW_ROWS = 500


def _excel_to_html_response(
    file_path: str,
    filename: str,
    active_sheet: int | None = None,
    token: str | None = None,
):
    """Convert an Excel file to a styled HTML page for inline viewing.

    Only renders the active sheet (default: first). Truncates at MAX_PREVIEW_ROWS
    to keep response times fast for large spreadsheets. Sheet tabs let users
    switch between sheets via server-side reload (?sheet=N).
    """
    from fastapi.responses import HTMLResponse
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {e}")

    sheet_names = wb.sheetnames
    active_idx = active_sheet if active_sheet is not None and 0 <= active_sheet < len(sheet_names) else 0

    html_parts = [
        "<!DOCTYPE html><html><head>",
        "<meta charset='utf-8'>",
        f"<title>{escape(filename)}</title>",
        "<style>",
        "  * { box-sizing: border-box; margin: 0; padding: 0; }",
        "  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f9fafb; color: #111827; padding: 24px; }",
        "  h1 { font-size: 16px; font-weight: 600; margin-bottom: 16px; color: #374151; }",
        "  h2 { font-size: 14px; font-weight: 600; margin: 8px 0; color: #6b7280; }",
        "  table { border-collapse: collapse; width: 100%; margin-bottom: 16px; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }",
        "  th { background: #f3f4f6; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em; color: #6b7280; position: sticky; top: 0; z-index: 1; }",
        "  th, td { padding: 8px 12px; border: 1px solid #e5e7eb; text-align: left; font-size: 13px; white-space: nowrap; }",
        "  tr:nth-child(even) { background: #f9fafb; }",
        "  tr:hover { background: #eff6ff; }",
        "  td.num { text-align: right; font-variant-numeric: tabular-nums; }",
        "  .sheet-tabs { display: flex; gap: 4px; margin-bottom: 16px; flex-wrap: wrap; }",
        "  .sheet-tab { padding: 6px 12px; font-size: 12px; border-radius: 6px; background: #e5e7eb; cursor: pointer; border: none; color: #374151; text-decoration: none; }",
        "  .sheet-tab.active { background: #3b82f6; color: white; }",
        "  .truncated { padding: 12px; text-align: center; color: #9ca3af; font-size: 13px; font-style: italic; }",
        "</style>",
        "</head><body>",
        f"<h1>{escape(filename)}</h1>",
    ]

    # Sheet tabs — each is a link that reloads with ?sheet=N (server-side, no full Excel re-parse overhead for client)
    if len(sheet_names) > 1:
        html_parts.append("<div class='sheet-tabs'>")
        for i, name in enumerate(sheet_names):
            active_cls = "active" if i == active_idx else ""
            # Build URL that swaps the sheet param
            query = {"sheet": i}
            if token:
                query["token"] = token
            html_parts.append(
                f"<a class='sheet-tab {active_cls}' href='?{urlencode(query)}'>{escape(name)}</a>"
            )
        html_parts.append("</div>")

    # Render only the active sheet
    ws = wb[sheet_names[active_idx]]
    html_parts.append(f"<h2>{escape(sheet_names[active_idx])}</h2>")
    html_parts.append("<div style='overflow-x:auto'><table>")

    row_count = 0
    truncated = False
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx >= MAX_PREVIEW_ROWS:
            truncated = True
            break
        tag = "th" if row_idx == 0 else "td"
        html_parts.append("<tr>")
        for cell in row:
            val = "" if cell is None else str(cell)
            # Escape HTML entities for all cells (headers and data)
            val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            cls = ""
            if tag == "td":
                try:
                    float(str(cell).replace(",", "").replace("$", "").replace("%", ""))
                    cls = " class='num'"
                except (ValueError, TypeError, AttributeError):
                    pass
            html_parts.append(f"<{tag}{cls}>{val}</{tag}>")
        html_parts.append("</tr>")
        row_count += 1

    html_parts.append("</table></div>")

    if truncated:
        html_parts.append(
            f"<div class='truncated'>Showing first {MAX_PREVIEW_ROWS} of {ws.max_row or '?'} rows. "
            f"Download the file for the complete data.</div>"
        )

    html_parts.append(f"<div style='margin-top:8px;font-size:12px;color:#9ca3af;'>{row_count} rows displayed</div>")
    html_parts.append("</body></html>")

    wb.close()
    return HTMLResponse(
        content="\n".join(html_parts),
        headers={
            # Sheet-name/cell escaping is the primary defense; the CSP is
            # belt-and-suspenders so a missed sink still can't run script.
            "Content-Security-Policy": "default-src 'none'; style-src 'unsafe-inline'",
            "X-Content-Type-Options": "nosniff",
        },
    )
