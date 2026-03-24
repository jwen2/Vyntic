"""
Generate extended due diligence documents for Acme Cloud Solutions.
Creates: QoE Report, LBO Model, SIM, Legal Summary, Operational Review.
"""
import os
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
except ImportError:
    os.system("pip install reportlab")
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(__file__).parent

# ============================================================================
# Shared styling
# ============================================================================

def get_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DealTitle", parent=styles["Title"],
                                  fontSize=18, textColor=colors.HexColor("#1a365d"))
    h1 = ParagraphStyle("H1", parent=styles["Heading1"],
                         fontSize=14, textColor=colors.HexColor("#2d3748"),
                         spaceBefore=18, spaceAfter=8)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"],
                         fontSize=11, textColor=colors.HexColor("#4a5568"),
                         spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle("Body", parent=styles["Normal"],
                           fontSize=9.5, leading=14, spaceBefore=4, spaceAfter=4)
    bullet = ParagraphStyle("Bullet", parent=body, leftIndent=20,
                             bulletIndent=10, spaceBefore=2, spaceAfter=2)
    conf = ParagraphStyle("Conf", parent=styles["Normal"], fontSize=8,
                           textColor=colors.red, alignment=1)
    return title_style, h1, h2, body, bullet, conf


def make_table(headers, rows, col_widths=None):
    data = [headers] + rows
    if col_widths is None:
        col_widths = [1.8 * inch] + [1.1 * inch] * (len(headers) - 1)
        total = sum(col_widths)
        if total > 6.5 * inch:
            ratio = 6.5 * inch / total
            col_widths = [w * ratio for w in col_widths]
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return t


# ============================================================================
# 1. Quality of Earnings Report (QoE)
# ============================================================================

def generate_qoe_report():
    filename = OUT / "acme_saas_qoe_report.pdf"
    doc = SimpleDocTemplate(str(filename), pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    title_style, h1, h2, body, bullet, conf = get_styles()
    elements = []

    # Title page
    elements.append(Spacer(1, 2 * inch))
    elements.append(Paragraph("CONFIDENTIAL", conf))
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Quality of Earnings Report", title_style))
    elements.append(Paragraph("Acme Cloud Solutions, Inc.", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph("Prepared by: Apex Advisory Partners, LLC", body))
    elements.append(Paragraph("Date: January 15, 2025", body))
    elements.append(Paragraph("Period Covered: FY2022 - FY2024 (TTM through Q3 2024)", body))
    elements.append(PageBreak())

    # Executive Summary
    elements.append(Paragraph("1. Executive Summary", h1))
    elements.append(Paragraph(
        "Apex Advisory Partners was engaged to perform a quality of earnings analysis on "
        "Acme Cloud Solutions, Inc. (the 'Company') in connection with a potential acquisition. "
        "Our analysis covers FY2022 through FY2024 with a focus on earnings quality, revenue "
        "sustainability, and the appropriateness of management's EBITDA adjustments. Overall, "
        "we find the Company's reported EBITDA of $5.9M (14.0% margin) for FY2024 to be "
        "materially accurate, with net downward adjustments of $0.4M resulting in Adjusted "
        "EBITDA of $5.5M (13.1% margin).", body))

    elements.append(Paragraph("2. EBITDA Bridge & Adjustments", h1))
    elements.append(Paragraph(
        "Management has presented Adjusted EBITDA of $5.9M for FY2024. Our independent analysis "
        "identifies the following adjustments:", body))

    elements.append(make_table(
        ["Adjustment Item", "Mgmt Adj ($K)", "Apex Adj ($K)", "Variance ($K)", "Comment"],
        [
            ["Reported EBITDA", "5,880", "5,880", "—", "Per audited financials"],
            ["Stock-based compensation", "+820", "+820", "—", "Non-cash; accepted"],
            ["One-time legal settlement", "+340", "+340", "—", "Verified; non-recurring"],
            ["Founder bonus (discretionary)", "+280", "+0", "-280", "Recurring annual payment"],
            ["Office relocation costs", "+150", "+150", "—", "Non-recurring; verified"],
            ["Customer migration credits", "+0", "-120", "-120", "Revenue reduction; recurring"],
            ["Mgmt Adjusted EBITDA", "7,470", "7,070", "-400", ""],
            ["Apex Adjusted EBITDA", "", "5,500", "", "After reversing non-qualifying items"],
        ],
    ))

    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(
        "Key Finding: The $280K 'founder bonus' classified as discretionary is in fact a contractual "
        "annual payment to the CEO and CTO per their employment agreements (Section 4.2). This payment "
        "has been made consistently for 5 years and should be treated as a recurring operating expense. "
        "Additionally, customer migration credits of $120K annually represent ongoing platform upgrade "
        "costs that reduce net revenue.", body))

    elements.append(PageBreak())

    # Revenue Quality
    elements.append(Paragraph("3. Revenue Quality Analysis", h1))
    elements.append(Paragraph(
        "Revenue quality is strong overall, with 92% of total revenue classified as recurring "
        "subscription revenue. Key observations:", body))

    elements.append(make_table(
        ["Revenue Category", "FY2022 ($K)", "FY2023 ($K)", "FY2024 ($K)", "% of FY2024"],
        [
            ["Subscription - Annual", "17,360", "24,480", "33,880", "80.7%"],
            ["Subscription - Monthly", "4,340", "4,320", "4,720", "11.2%"],
            ["Professional Services", "2,400", "2,700", "3,400", "8.1%"],
            ["Total Revenue", "24,100", "31,500", "42,000", "100.0%"],
        ],
    ))

    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph("Revenue Quality Flags:", h2))
    elements.append(Paragraph("- Monthly subscription revenue ($4.7M) carries higher churn risk than annual contracts. "
                              "Monthly subscriber churn is 2.8% per month vs. 4.2% annual churn for annual contracts.", bullet))
    elements.append(Paragraph("- Professional services revenue is non-recurring but has grown at 19% CAGR, suggesting "
                              "deepening customer relationships rather than one-time implementation.", bullet))
    elements.append(Paragraph("- Three large contracts ($400K+ each) were signed in Q4 2024 representing $1.8M of "
                              "ARR. Two include 90-day termination-for-convenience clauses.", bullet))
    elements.append(Paragraph("- Net revenue retention of 118% is verified; however, excluding the top 10 accounts "
                              "by expansion, NRR drops to 109%.", bullet))

    elements.append(PageBreak())

    # Working Capital
    elements.append(Paragraph("4. Working Capital Analysis", h1))
    elements.append(Paragraph(
        "Working capital is favorable for the business model. Deferred revenue (prepaid subscriptions) "
        "provides a natural source of working capital. Key metrics:", body))

    elements.append(make_table(
        ["Metric", "FY2022", "FY2023", "FY2024", "Trend"],
        [
            ["DSO (Days Sales Outstanding)", "62 days", "60 days", "59 days", "Improving"],
            ["Deferred Revenue ($M)", "$6.4M", "$8.5M", "$10.5M", "Growing"],
            ["Prepaid Expenses ($M)", "$0.85M", "$1.1M", "$1.4M", "Growing"],
            ["Net Working Capital ($M)", "$2.85M", "$5.3M", "$10.1M", "Strong"],
            ["NWC as % of Revenue", "11.8%", "16.8%", "24.0%", "Expanding"],
            ["Cash Conversion Cycle", "-22 days", "-25 days", "-28 days", "Favorable"],
        ],
    ))

    elements.append(Paragraph(
        "Finding: The negative cash conversion cycle (-28 days) is a structural advantage of the SaaS "
        "model — customers pay upfront while revenue is recognized ratably. This provides the Company "
        "with an embedded source of financing that reduces external capital needs.", body))

    elements.append(PageBreak())

    # Customer Cohort Analysis
    elements.append(Paragraph("5. Customer Cohort Analysis", h1))
    elements.append(Paragraph(
        "We analyzed revenue retention by annual cohort to assess sustainability of the Company's "
        "stated 118% net revenue retention:", body))

    elements.append(make_table(
        ["Cohort Year", "Starting ARR ($K)", "Year 1 NRR", "Year 2 NRR", "Year 3 NRR", "Churned Logos"],
        [
            ["2019 Cohort", "3,200", "108%", "112%", "115%", "12 of 85 (14%)"],
            ["2020 Cohort", "4,800", "110%", "114%", "118%", "8 of 92 (9%)"],
            ["2021 Cohort", "6,100", "112%", "116%", "—", "6 of 105 (6%)"],
            ["2022 Cohort", "8,500", "115%", "120%", "—", "5 of 120 (4%)"],
            ["2023 Cohort", "10,200", "118%", "—", "—", "3 of 135 (2%)"],
        ],
    ))

    elements.append(Paragraph(
        "Finding: Cohort analysis confirms improving retention trends. Newer cohorts show higher NRR "
        "and lower logo churn, consistent with product maturation and stronger onboarding processes. "
        "The 2019 cohort shows elevated churn (14% logo churn) which management attributes to early "
        "product limitations that have since been addressed. We recommend monitoring the 2023 cohort "
        "closely as it matures — current NRR of 118% in Year 1 is above historical averages and may "
        "normalize downward.", body))

    elements.append(PageBreak())

    # Tax and Compliance
    elements.append(Paragraph("6. Tax Position & Compliance", h1))
    elements.append(Paragraph(
        "The Company has federal net operating loss (NOL) carryforwards of approximately $8.2M as of "
        "FY2024, primarily generated during the growth investment phase (FY2019-FY2022). Key tax observations:", body))
    elements.append(Paragraph("- NOL carryforwards of $8.2M are subject to Section 382 limitations upon change "
                              "of control. Estimated annual limitation: $1.6M based on current long-term tax-exempt rate.", bullet))
    elements.append(Paragraph("- R&D tax credits of $1.4M have been claimed but not yet audited by the IRS. "
                              "Management estimates $1.1M would survive an audit based on qualified expenditure documentation.", bullet))
    elements.append(Paragraph("- The Company has nexus in 14 states but only files in 11. Three states "
                              "(New York, Illinois, New Jersey) represent potential exposure of $340K in back taxes plus penalties.", bullet))
    elements.append(Paragraph("- Transfer pricing: No intercompany transactions; all operations are domestic.", bullet))

    doc.build(elements)
    print(f"  Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# 2. LBO Model / Investment Returns Analysis
# ============================================================================

def generate_lbo_model():
    filename = OUT / "acme_saas_lbo_model.xlsx"
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    section_font = Font(name="Calibri", bold=True, size=11, color="1A365D")
    label_font = Font(name="Calibri", bold=True, size=10)
    number_font = Font(name="Calibri", size=10)
    pct_font = Font(name="Calibri", size=10, italic=True, color="4A5568")
    highlight_fill = PatternFill(start_color="EBF5FF", end_color="EBF5FF", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="CBD5E0"))

    # --- Transaction Summary Sheet ---
    ws = wb.active
    ws.title = "Transaction Summary"
    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="Acme Cloud Solutions — LBO Transaction Summary").font = Font(
        name="Calibri", bold=True, size=14, color="1A365D")

    tx_data = [
        ["", ""],
        ["TRANSACTION ASSUMPTIONS", ""],
        ["Enterprise Value ($M)", 252.0],
        ["EV / FY2024 Revenue", "6.0x"],
        ["EV / FY2024 EBITDA", "42.7x"],
        ["EV / FY2025E EBITDA", "23.1x"],
        ["EV / FY2024 ARR", "6.0x"],
        ["", ""],
        ["SOURCES & USES ($M)", ""],
        ["Sponsor Equity", 152.0],
        ["Senior Term Loan (4.5x EBITDA)", 65.0],
        ["Revolving Credit Facility", 10.0],
        ["Seller Rollover Equity (15%)", 25.0],
        ["Total Sources", 252.0],
        ["", ""],
        ["Enterprise Value", 252.0],
        ["Transaction Fees (2.5%)", 6.3],
        ["Debt Financing Fees (1.5%)", 1.1],
        ["Total Uses", 259.4],
        ["", ""],
        ["DEBT TERMS", ""],
        ["Senior Term Loan Rate", "SOFR + 425bps (est. 9.5%)"],
        ["Revolver Rate", "SOFR + 375bps (est. 9.0%)"],
        ["Term Loan Maturity", "7 years"],
        ["Mandatory Amortization", "1% annual"],
        ["Leverage at Close", "2.8x Net Debt / FY2025E EBITDA"],
    ]

    for r, (label, val) in enumerate(tx_data, 3):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=val)
        c1.border = thin_border
        c2.border = thin_border
        if label and label == label.upper() and label.replace(" ", "").replace("(", "").replace(")", "").replace("$", "").replace("&", "").replace("/", "").isalpha():
            c1.font = section_font
        elif label in ("Total Sources", "Total Uses"):
            c1.font = Font(name="Calibri", bold=True, size=10, color="1A365D")
            c2.font = Font(name="Calibri", bold=True, size=10, color="1A365D")
            for col in [1, 2]:
                ws.cell(row=r, column=col).fill = highlight_fill
        else:
            c1.font = label_font
            if isinstance(val, (int, float)):
                c2.font = number_font
                c2.number_format = "#,##0.0"
            else:
                c2.font = pct_font
        c2.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30

    # --- Projected Returns Sheet ---
    ws2 = wb.create_sheet("Projected Returns")
    ws2.merge_cells("A1:G1")
    ws2.cell(row=1, column=1, value="Acme Cloud Solutions — 5-Year LBO Returns Analysis").font = Font(
        name="Calibri", bold=True, size=14, color="1A365D")

    headers = ["Metric", "FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    proj_data = [
        ["Revenue ($M)", 54.6, 68.3, 83.3, 99.9, 117.9],
        ["Revenue Growth (%)", "30%", "25%", "22%", "20%", "18%"],
        ["Gross Profit ($M)", 46.9, 59.1, 72.5, 87.9, 104.8],
        ["Gross Margin (%)", "86%", "87%", "87%", "88%", "89%"],
        ["EBITDA ($M)", 10.9, 15.4, 20.8, 27.0, 34.5],
        ["EBITDA Margin (%)", "20%", "23%", "25%", "27%", "29%"],
        ["Capex ($M)", 2.7, 3.4, 4.2, 5.0, 5.9],
        ["Capex as % of Revenue", "5%", "5%", "5%", "5%", "5%"],
        ["Free Cash Flow ($M)", 7.8, 11.2, 15.6, 20.8, 27.1],
        ["FCF Conversion (%)", "72%", "73%", "75%", "77%", "79%"],
        ["", "", "", "", "", ""],
        ["Debt Balance ($M)", 74.4, 63.2, 47.6, 26.8, 0.0],
        ["Net Debt ($M)", 56.4, 34.0, 14.0, -12.0, -45.1],
        ["Net Leverage (x)", "5.2x", "2.2x", "0.7x", "-0.4x", "-1.3x"],
        ["", "", "", "", "", ""],
        ["EXIT ANALYSIS (FY2029E)", "", "", "", "", ""],
        ["Exit Multiple (EV/EBITDA)", "", "", "", "", ""],
        ["  Base Case: 15.0x", "", "", "", "", 517.5],
        ["  Upside Case: 18.0x", "", "", "", "", 621.0],
        ["  Downside Case: 12.0x", "", "", "", "", 414.0],
        ["", "", "", "", "", ""],
        ["Equity Value at Exit ($M)", "", "", "", "", ""],
        ["  Base Case", "", "", "", "", 562.6],
        ["  Upside Case", "", "", "", "", 666.1],
        ["  Downside Case", "", "", "", "", 459.1],
        ["", "", "", "", "", ""],
        ["Sponsor IRR", "", "", "", "", ""],
        ["  Base Case", "", "", "", "", "30.0%"],
        ["  Upside Case", "", "", "", "", "34.5%"],
        ["  Downside Case", "", "", "", "", "24.8%"],
        ["", "", "", "", "", ""],
        ["MOIC", "", "", "", "", ""],
        ["  Base Case", "", "", "", "", "3.7x"],
        ["  Upside Case", "", "", "", "", "4.4x"],
        ["  Downside Case", "", "", "", "", "3.0x"],
    ]

    for r, row in enumerate(proj_data, 4):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = label_font if val and not str(val).startswith("  ") else number_font
                if val and val == val.upper() and not val.startswith(" "):
                    cell.font = section_font
            elif isinstance(val, (int, float)) and val != "":
                cell.font = number_font
                cell.number_format = "#,##0.0"
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, str) and ("%" in val or "x" in val):
                cell.font = pct_font
                cell.alignment = Alignment(horizontal="right")

    ws2.column_dimensions["A"].width = 30
    for col in range(2, 7):
        ws2.column_dimensions[chr(64 + col)].width = 14

    wb.save(str(filename))
    print(f"  Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# 3. Seller's Information Memorandum (SIM)
# ============================================================================

def generate_sim():
    filename = OUT / "acme_saas_sim.pdf"
    doc = SimpleDocTemplate(str(filename), pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    title_style, h1, h2, body, bullet, conf = get_styles()
    elements = []

    # Title page
    elements.append(Spacer(1, 2 * inch))
    elements.append(Paragraph("STRICTLY CONFIDENTIAL", conf))
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Seller's Information Memorandum", title_style))
    elements.append(Paragraph("Acme Cloud Solutions, Inc.", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph("Prepared by: Meridian Capital Advisors", body))
    elements.append(Paragraph("Project Atlas | February 2025", body))
    elements.append(PageBreak())

    # Company Overview
    elements.append(Paragraph("1. Company Overview", h1))
    elements.append(Paragraph(
        "Acme Cloud Solutions ('Acme' or the 'Company') is a leading provider of cloud-based "
        "Enterprise Resource Planning (ERP) software purpose-built for mid-market manufacturing "
        "and distribution companies. Founded in 2014 and headquartered in Austin, TX, the Company "
        "has grown to serve over 450 enterprise customers across North America, generating $42M in "
        "Annual Recurring Revenue (ARR) with best-in-class unit economics.", body))

    elements.append(Paragraph("2. Market Opportunity", h1))
    elements.append(Paragraph(
        "The mid-market ERP software market represents an $18 billion total addressable market (TAM), "
        "growing at approximately 12% annually. The market is experiencing a secular shift from "
        "on-premise to cloud-native solutions, with cloud penetration currently at only 35% in "
        "the mid-market segment. Key market dynamics:", body))
    elements.append(Paragraph("- $18B TAM for mid-market ERP (companies with $50M-$1B revenue)", bullet))
    elements.append(Paragraph("- $6.3B serviceable addressable market (SAM) — manufacturing & distribution vertical", bullet))
    elements.append(Paragraph("- $1.8B serviceable obtainable market (SOM) — cloud-native, North America", bullet))
    elements.append(Paragraph("- Current market share: approximately 2.3% of SOM", bullet))
    elements.append(Paragraph("- Cloud adoption CAGR of 18% in mid-market vs. 8% overall ERP market", bullet))

    elements.append(Paragraph("3. Competitive Landscape", h1))
    elements.append(make_table(
        ["Competitor", "Target Market", "Deployment", "Estimated Revenue", "Key Strength"],
        [
            ["SAP S/4HANA", "Enterprise", "Hybrid", "$30B+", "Incumbent; global scale"],
            ["Oracle Fusion", "Enterprise", "Cloud", "$12B+", "Database integration"],
            ["NetSuite (Oracle)", "SMB/Mid-Market", "Cloud", "$2.5B", "Broad functionality"],
            ["Epicor Kinetic", "Mid-Market Mfg", "Hybrid", "$1.0B", "Manufacturing depth"],
            ["Infor CloudSuite", "Mid-Market", "Cloud", "$3.3B", "Industry-specific"],
            ["Acme Cloud", "Mid-Market Mfg", "Cloud-native", "$42M", "Modern UX; fast deploy"],
        ],
    ))

    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph(
        "Acme's competitive advantage lies in its cloud-native architecture (vs. legacy lift-and-shift), "
        "rapid 90-day implementation timeline (vs. 12-18 months for SAP/Oracle), and purpose-built "
        "manufacturing workflows. Customer win rates against incumbents have improved from 22% in 2022 "
        "to 38% in 2024.", body))

    elements.append(PageBreak())

    # Management Team
    elements.append(Paragraph("4. Management Team", h1))
    elements.append(make_table(
        ["Name", "Title", "Tenure", "Background", "Equity %"],
        [
            ["David Chen", "CEO & Co-Founder", "10 years", "VP Product at SAP; sold prior SaaS co for $85M", "18%"],
            ["Sarah Kim", "CTO & Co-Founder", "10 years", "Sr. Director Engineering at Oracle; MIT CS", "14%"],
            ["Marcus Rivera", "CFO", "3 years", "CFO at 2 PE-backed SaaS exits ($200M+, $450M)", "4%"],
            ["Lisa Park", "VP Sales", "4 years", "Regional VP at Salesforce; built $50M ARR team", "3%"],
            ["James Wright", "VP Customer Success", "5 years", "VP CS at HubSpot; reduced churn by 40%", "2%"],
            ["Rachel Foster", "VP Engineering", "2 years", "Staff Engineer at Google Cloud; 15 yrs exp", "2%"],
        ],
    ))

    elements.append(Paragraph(
        "Key Person Risk: CTO Sarah Kim holds critical architectural knowledge of the platform's "
        "multi-tenant infrastructure. While VP Engineering Rachel Foster has been onboarded to share "
        "this knowledge, a transition plan has not been formally documented. Both founders have "
        "indicated willingness to enter 2-year post-close employment agreements with appropriate "
        "incentive structures.", body))

    # Go-to-Market
    elements.append(Paragraph("5. Go-to-Market Strategy", h1))
    elements.append(Paragraph(
        "The Company employs a hybrid direct sales and channel partner model:", body))
    elements.append(Paragraph("- Direct sales team: 22 AEs generating ~$1.9M ARR per rep", bullet))
    elements.append(Paragraph("- Channel partners: 8 certified implementation partners driving 25% of new logos", bullet))
    elements.append(Paragraph("- Inbound marketing: 45% of qualified pipeline from content/SEO/events", bullet))
    elements.append(Paragraph("- Average deal size: $93K ACV (up from $72K in FY2022)", bullet))
    elements.append(Paragraph("- Sales cycle: 4.2 months average (down from 5.8 months in FY2022)", bullet))
    elements.append(Paragraph("- Win rate: 38% on competitive deals (up from 22% in FY2022)", bullet))
    elements.append(Paragraph("- CAC: $62K blended ($85K direct, $38K channel)", bullet))
    elements.append(Paragraph("- LTV/CAC: 5.2x (industry benchmark: 3.0x+)", bullet))

    elements.append(PageBreak())

    # Growth Levers
    elements.append(Paragraph("6. Value Creation Opportunities", h1))
    elements.append(Paragraph("Identified growth levers for a new ownership platform:", body))
    elements.append(Paragraph("- Product: AI-powered demand forecasting module (beta Q2 2025, $8K ACV uplift)", bullet))
    elements.append(Paragraph("- Geographic: European expansion ($4B addressable market, zero presence today)", bullet))
    elements.append(Paragraph("- Pricing: Usage-based pricing tier for high-volume customers (est. +15% ARPU)", bullet))
    elements.append(Paragraph("- M&A: 3 identified bolt-on targets in adjacent verticals (food & bev, chemicals)", bullet))
    elements.append(Paragraph("- Channel: Expand from 8 to 20+ implementation partners by FY2027", bullet))
    elements.append(Paragraph("- Product: Mobile-first field operations module (customer #1 request)", bullet))

    # Customer Testimonials
    elements.append(Paragraph("7. Customer Case Studies", h1))

    elements.append(Paragraph("Midwest Precision Parts ($180M revenue, 800 employees)", h2))
    elements.append(Paragraph(
        "Challenge: Replaced legacy on-premise ERP (Epicor Vantage) that required 3 FTEs to maintain. "
        "Solution: Implemented Acme's full platform in 82 days. Results: Reduced order-to-cash cycle "
        "from 45 days to 28 days, eliminated $240K/year in IT maintenance costs, achieved 99.97% "
        "system uptime. Annual contract value: $210K. Customer since 2021.", body))

    elements.append(Paragraph("Pacific Coast Distributors ($95M revenue, 350 employees)", h2))
    elements.append(Paragraph(
        "Challenge: Managing inventory across 6 warehouses with spreadsheets and a legacy system. "
        "Solution: Deployed Acme's supply chain and inventory modules. Results: Reduced stockouts by "
        "62%, improved inventory turnover from 4.2x to 6.8x, saving $1.2M annually in carrying costs. "
        "Annual contract value: $145K. Customer since 2022. Expanded to workforce module in 2024.", body))

    doc.build(elements)
    print(f"  Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# 4. Legal Due Diligence Summary
# ============================================================================

def generate_legal_summary():
    filename = OUT / "acme_saas_legal_dd_summary.pdf"
    doc = SimpleDocTemplate(str(filename), pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    title_style, h1, h2, body, bullet, conf = get_styles()
    elements = []

    # Title page
    elements.append(Spacer(1, 2 * inch))
    elements.append(Paragraph("ATTORNEY-CLIENT PRIVILEGED", conf))
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Legal Due Diligence Summary", title_style))
    elements.append(Paragraph("Acme Cloud Solutions, Inc.", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph("Prepared by: Sterling & Associates LLP", body))
    elements.append(Paragraph("Date: February 1, 2025", body))
    elements.append(PageBreak())

    # Corporate Structure
    elements.append(Paragraph("1. Corporate Structure & Governance", h1))
    elements.append(Paragraph(
        "Acme Cloud Solutions, Inc. is a Delaware C-Corporation incorporated on March 12, 2014. "
        "The Company has two classes of stock: Common Stock (10,000,000 shares authorized, 6,800,000 "
        "issued) and Series A Preferred Stock (2,000,000 shares authorized, 1,500,000 issued). "
        "Preferred holders have standard protective provisions including board seat rights, "
        "anti-dilution protection, and liquidation preference of 1x non-participating.", body))
    elements.append(Paragraph(
        "Board Composition: 5 directors — 2 founder seats, 2 investor seats (Highland Ventures), "
        "1 independent director (appointed 2023). Quorum requires 3 directors. Change of control "
        "requires approval of 2/3 of common shares and majority of preferred shares.", body))

    elements.append(Paragraph("2. Material Contracts", h1))
    elements.append(Paragraph("The following contracts are considered material:", body))

    elements.append(make_table(
        ["Contract", "Counterparty", "Annual Value", "Term", "CoC Provision"],
        [
            ["Master Sub Agreement", "AWS (hosting)", "$3.4M", "3yr (exp 2026)", "None; standard ToS"],
            ["Enterprise License", "Microsoft 365", "$420K", "Annual", "None"],
            ["Channel Partner", "Deloitte Consulting", "$2.1M sourced", "2yr (exp 2025)", "30-day notice"],
            ["Office Lease", "WeWork Austin", "$680K", "3yr (exp 2027)", "Assignable with consent"],
            ["Credit Facility", "Silicon Valley Bank", "$10M revolver", "2yr (exp 2025)", "Acceleration clause"],
            ["IP License", "Stanford University", "$85K/yr", "Perpetual", "None"],
        ],
    ))

    elements.append(PageBreak())

    # Litigation
    elements.append(Paragraph("3. Litigation & Disputes", h1))
    elements.append(Paragraph("3.1 Active Litigation", h2))
    elements.append(Paragraph(
        "Acme Cloud Solutions v. DataSync Corp. (Case No. 2024-CV-03847, Travis County District Court): "
        "Patent infringement claim filed by Acme in August 2024 alleging DataSync's 'CloudBridge' "
        "product infringes U.S. Patent No. 10,892,441 covering Acme's real-time data synchronization "
        "method. Case is in discovery phase. Estimated legal costs through trial: $800K-$1.2M. "
        "Management believes the claim is strong based on prior art analysis.", body))

    elements.append(Paragraph("3.2 Settled Matters", h2))
    elements.append(Paragraph(
        "Former Employee Wrongful Termination (Jane Doe v. Acme, settled December 2023): "
        "Former VP of Marketing alleged wrongful termination and breach of employment agreement. "
        "Settled for $340K (included in FY2024 one-time adjustments). Release and non-disparagement "
        "agreement executed. No admission of liability.", body))

    elements.append(Paragraph("3.3 Threatened Claims", h2))
    elements.append(Paragraph(
        "Demand letter received from Oracle Corporation (November 2024) alleging potential trade "
        "secret misappropriation related to two engineers hired from Oracle's cloud division in 2023. "
        "Outside counsel has reviewed and believes the claim lacks merit — both employees had standard "
        "(not enhanced) non-compete agreements that expired prior to joining Acme. No lawsuit has been "
        "filed. Estimated contingent liability: $0 (low probability of adverse outcome).", body))

    # IP
    elements.append(Paragraph("4. Intellectual Property", h1))
    elements.append(Paragraph(
        "IP Portfolio Summary:", body))
    elements.append(Paragraph("- 6 issued U.S. utility patents (data sync, multi-tenant architecture, predictive scheduling)", bullet))
    elements.append(Paragraph("- 3 pending patent applications (AI forecasting, workflow automation, edge computing)", bullet))
    elements.append(Paragraph("- 12 registered trademarks (Acme Cloud, CloudERP, AcmePlan, etc.)", bullet))
    elements.append(Paragraph("- All IP is owned by the Company (no individual inventor assignments outstanding)", bullet))
    elements.append(Paragraph("- Stanford University perpetual license for foundational ML algorithms ($85K/yr royalty)", bullet))
    elements.append(Paragraph("- Open source audit: 142 OSS components used; 8 flagged as copyleft (GPL) — "
                              "engineering team has isolated these from proprietary code. No contamination risk identified.", bullet))

    elements.append(PageBreak())

    # Data Privacy
    elements.append(Paragraph("5. Data Privacy & Cybersecurity", h1))
    elements.append(Paragraph(
        "Regulatory Compliance Status:", body))
    elements.append(Paragraph("- SOC 2 Type II certified (renewed August 2024, no material findings)", bullet))
    elements.append(Paragraph("- GDPR: Currently non-compliant (no EU customers; compliance required pre-expansion)", bullet))
    elements.append(Paragraph("- CCPA: Compliant; privacy policy updated Q1 2024", bullet))
    elements.append(Paragraph("- HIPAA: Not applicable (no healthcare data processed)", bullet))
    elements.append(Paragraph("- Data breach history: One incident in March 2023 — unauthorized access to staging "
                              "environment (non-production). No customer data exposed. Root cause: misconfigured IAM "
                              "role. Remediated within 4 hours. Incident reported to affected parties per policy.", bullet))
    elements.append(Paragraph("- Penetration testing: Annual third-party pen test by CrowdStrike. Most recent "
                              "(October 2024) identified 2 medium-severity findings, both remediated within 30 days.", bullet))
    elements.append(Paragraph("- Cyber insurance: $10M aggregate policy with Chubb (premium: $48K/yr)", bullet))

    # Employment
    elements.append(Paragraph("6. Employment & Labor Matters", h1))
    elements.append(Paragraph(
        "Key employment observations:", body))
    elements.append(Paragraph("- 320 total employees (280 full-time, 40 contractors)", bullet))
    elements.append(Paragraph("- No collective bargaining agreements or union representation", bullet))
    elements.append(Paragraph("- All employees have signed invention assignment and confidentiality agreements", bullet))
    elements.append(Paragraph("- Non-compete agreements: CEO, CTO, CFO, VP Sales (12-month post-termination, "
                              "enforceable in Texas per 2024 reform)", bullet))
    elements.append(Paragraph("- Stock option pool: 15% of fully diluted shares (2,400,000 options), "
                              "1,680,000 vested, 720,000 unvested", bullet))
    elements.append(Paragraph("- Change of control acceleration: 50% of unvested options accelerate on "
                              "single-trigger; remaining 50% on double-trigger (termination within 12 months)", bullet))
    elements.append(Paragraph("- Estimated cost of option acceleration at $252M EV: $18.2M (single-trigger) "
                              "plus up to $18.2M (double-trigger if applicable)", bullet))
    elements.append(Paragraph("- WARN Act: No facility closures or mass layoffs planned", bullet))
    elements.append(Paragraph("- I-9 Compliance: 28 employees on H-1B visas (engineering); all current", bullet))

    elements.append(PageBreak())

    # Regulatory
    elements.append(Paragraph("7. Regulatory & Environmental", h1))
    elements.append(Paragraph(
        "The Company operates in a lightly regulated industry (enterprise software). Key regulatory observations:", body))
    elements.append(Paragraph("- No industry-specific regulatory licenses required", bullet))
    elements.append(Paragraph("- Export controls: ITAR not applicable; EAR classification is EAR99 (unrestricted)", bullet))
    elements.append(Paragraph("- Anti-bribery/FCPA: No international operations; no government customers; "
                              "anti-corruption policy adopted 2022", bullet))
    elements.append(Paragraph("- Environmental: Office-only operations; no manufacturing. No environmental "
                              "liabilities identified.", bullet))
    elements.append(Paragraph("- Tax compliance: State nexus issue identified in 3 states (see QoE report Section 6)", bullet))

    doc.build(elements)
    print(f"  Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# 5. Detailed Financial Statements (3-Statement Model)
# ============================================================================

def generate_detailed_financials():
    filename = OUT / "acme_saas_detailed_financials.xlsx"
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    section_font = Font(name="Calibri", bold=True, size=11, color="1A365D")
    label_font = Font(name="Calibri", bold=True, size=10)
    number_font = Font(name="Calibri", size=10)
    pct_font = Font(name="Calibri", size=10, italic=True, color="4A5568")
    thin_border = Border(bottom=Side(style="thin", color="CBD5E0"))

    # --- Quarterly P&L ---
    ws = wb.active
    ws.title = "Quarterly P&L"
    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value="Acme Cloud Solutions — Quarterly Income Statement ($K)").font = Font(
        name="Calibri", bold=True, size=14, color="1A365D")

    headers = ["", "Q1'24", "Q2'24", "Q3'24", "Q4'24", "FY2024", "Q1'25E", "Q2'25E", "FY2025E"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    quarterly_data = [
        ["Subscription Revenue", 8800, 9400, 10000, 10400, 38600, 11200, 12100, 50700],
        ["Professional Services", 750, 850, 900, 900, 3400, 950, 1000, 3900],
        ["Total Revenue", 9550, 10250, 10900, 11300, 42000, 12150, 13100, 54600],
        ["", "", "", "", "", "", "", "", ""],
        ["Cost of Revenue", 1433, 1538, 1635, 1695, 6300, 1701, 1834, 7644],
        ["Gross Profit", 8118, 8713, 9265, 9605, 35700, 10449, 11266, 46956],
        ["Gross Margin %", "85.0%", "85.0%", "85.0%", "85.0%", "85.0%", "86.0%", "86.0%", "86.0%"],
        ["", "", "", "", "", "", "", "", ""],
        ["Sales & Marketing", 3338, 3575, 3808, 3980, 14700, 3888, 4192, 17472],
        ["Research & Development", 2388, 2563, 2725, 2825, 10500, 2673, 2883, 12012],
        ["General & Administrative", 1050, 1125, 1200, 1245, 4620, 1337, 1441, 6006],
        ["Total OpEx", 6776, 7263, 7733, 8050, 29820, 7898, 8516, 35490],
        ["", "", "", "", "", "", "", "", ""],
        ["EBITDA", 1342, 1450, 1532, 1555, 5880, 2551, 2750, 11466],
        ["EBITDA Margin %", "14.1%", "14.1%", "14.1%", "13.8%", "14.0%", "21.0%", "21.0%", "21.0%"],
        ["D&A", 191, 205, 218, 226, 840, 243, 262, 1092],
        ["EBIT", 1151, 1245, 1314, 1329, 5040, 2308, 2488, 10374],
    ]

    for r, row in enumerate(quarterly_data, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = label_font
            elif isinstance(val, str) and "%" in val:
                cell.font = pct_font
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, (int, float)) and val != "":
                cell.font = number_font
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 28
    for col in range(2, 10):
        ws.column_dimensions[chr(64 + col)].width = 12

    # --- Cash Flow Statement ---
    ws2 = wb.create_sheet("Cash Flow Statement")
    ws2.merge_cells("A1:E1")
    ws2.cell(row=1, column=1, value="Acme Cloud Solutions — Cash Flow Statement ($K)").font = Font(
        name="Calibri", bold=True, size=14, color="1A365D")

    headers2 = ["", "FY2022", "FY2023", "FY2024"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    cf_data = [
        ["OPERATING ACTIVITIES", "", "", ""],
        ["Net Income", -2771, 937, 3720],
        ["Depreciation & Amortization", 482, 630, 840],
        ["Stock-Based Compensation", 520, 680, 820],
        ["Changes in Working Capital:", "", "", ""],
        ["  Accounts Receivable", -820, -1100, -1600],
        ["  Prepaid Expenses", -170, -250, -300],
        ["  Accounts Payable", 360, 400, 600],
        ["  Deferred Revenue", 1280, 2100, 2000],
        ["  Accrued Expenses", 420, 700, 700],
        ["Net Cash from Operations", -699, 4097, 6780],
        ["", "", "", ""],
        ["INVESTING ACTIVITIES", "", "", ""],
        ["Capital Expenditures", -964, -1260, -1680],
        ["Capitalized Software Dev", -1200, -1400, -1600],
        ["Acquisition of Intangibles", -400, 0, 0],
        ["Net Cash from Investing", -2564, -2660, -3280],
        ["", "", "", ""],
        ["FINANCING ACTIVITIES", "", "", ""],
        ["Repayment of Debt", -400, -500, -500],
        ["Proceeds from Stock Options", 120, 180, 250],
        ["Net Cash from Financing", -280, -320, -250],
        ["", "", "", ""],
        ["Net Change in Cash", -3543, 1117, 3250],
        ["Beginning Cash", 11743, 8200, 12500],
        ["Ending Cash", 8200, 12500, 18700],
        ["", "", "", ""],
        ["FREE CASH FLOW", "", "", ""],
        ["Cash from Operations", -699, 4097, 6780],
        ["Less: Capex", -964, -1260, -1680],
        ["Less: Capitalized Software", -1200, -1400, -1600],
        ["Free Cash Flow", -2863, 1437, 3500],
        ["FCF Margin", "-11.9%", "4.6%", "8.3%"],
        ["FCF per Share ($)", "-$0.42", "$0.21", "$0.51"],
    ]

    for r, row in enumerate(cf_data, 4):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                if val and val == val.upper() and val.replace(" ", "").replace(":", "").isalpha():
                    cell.font = section_font
                else:
                    cell.font = label_font
            elif isinstance(val, str) and ("%" in val or "$" in val):
                cell.font = pct_font
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, (int, float)) and val != "":
                cell.font = number_font
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    ws2.column_dimensions["A"].width = 30
    for col in range(2, 5):
        ws2.column_dimensions[chr(64 + col)].width = 14

    # --- SaaS Metrics Dashboard ---
    ws3 = wb.create_sheet("SaaS Metrics")
    ws3.merge_cells("A1:E1")
    ws3.cell(row=1, column=1, value="Acme Cloud Solutions — SaaS Operating Metrics").font = Font(
        name="Calibri", bold=True, size=14, color="1A365D")

    headers3 = ["Metric", "FY2022", "FY2023", "FY2024", "FY2025E"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    metrics_data = [
        ["GROWTH METRICS", "", "", "", ""],
        ["ARR ($M)", 26.3, 34.0, 42.0, 54.6],
        ["ARR Growth (%)", "37%", "29%", "24%", "30%"],
        ["New ARR ($M)", 7.1, 9.8, 11.2, 16.2],
        ["Expansion ARR ($M)", 2.8, 4.2, 5.8, 7.4],
        ["Churned ARR ($M)", -1.6, -2.1, -2.4, -2.6],
        ["Net New ARR ($M)", 8.3, 11.9, 14.6, 21.0],
        ["", "", "", "", ""],
        ["UNIT ECONOMICS", "", "", "", ""],
        ["Customers (end of period)", 310, 370, 450, 540],
        ["Net New Customers", 65, 60, 80, 90],
        ["Gross Logo Churn (%)", "8.2%", "6.5%", "5.1%", "4.5%"],
        ["Net Revenue Retention (%)", "112%", "115%", "118%", "120%"],
        ["Gross Revenue Retention (%)", "92%", "94%", "95%", "96%"],
        ["Average ACV ($K)", 78, 85, 93, 101],
        ["LTV ($K)", 468, 510, 558, 606],
        ["CAC ($K)", 72, 68, 62, 58],
        ["LTV/CAC Ratio", "6.5x", "7.5x", "9.0x", "10.4x"],
        ["CAC Payback (months)", 18, 16, 14, 13],
        ["", "", "", "", ""],
        ["EFFICIENCY METRICS", "", "", "", ""],
        ["Rule of 40 Score", "29%", "37%", "47%", "50%"],
        ["Burn Multiple", "0.5x", "0.2x", "-0.3x", "-0.5x"],
        ["Magic Number", "0.62", "0.78", "0.88", "0.95"],
        ["Revenue per Employee ($K)", 78, 90, 131, 156],
        ["ARR per Employee ($K)", 85, 97, 131, 156],
        ["Gross Margin (%)", "82%", "83%", "85%", "86%"],
        ["S&M as % of Revenue", "45%", "40%", "35%", "32%"],
        ["R&D as % of Revenue", "30%", "26%", "25%", "22%"],
        ["G&A as % of Revenue", "16%", "11%", "11%", "11%"],
    ]

    for r, row in enumerate(metrics_data, 4):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                if val and val == val.upper() and val.replace(" ", "").isalpha():
                    cell.font = section_font
                else:
                    cell.font = label_font
            elif isinstance(val, str) and ("%" in val or "x" in val):
                cell.font = pct_font
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, (int, float)) and val != "":
                cell.font = number_font
                cell.number_format = "#,##0.0" if isinstance(val, float) else "#,##0"
                cell.alignment = Alignment(horizontal="right")

    ws3.column_dimensions["A"].width = 30
    for col in range(2, 6):
        ws3.column_dimensions[chr(64 + col)].width = 14

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(filename))
    print(f"  Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# 6. Operational Review
# ============================================================================

def generate_operational_review():
    filename = OUT / "acme_saas_operational_review.pdf"
    doc = SimpleDocTemplate(str(filename), pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    title_style, h1, h2, body, bullet, conf = get_styles()
    elements = []

    elements.append(Spacer(1, 2 * inch))
    elements.append(Paragraph("CONFIDENTIAL", conf))
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Operational Due Diligence Review", title_style))
    elements.append(Paragraph("Acme Cloud Solutions, Inc.", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph("Prepared by: Apex Advisory Partners, LLC", body))
    elements.append(Paragraph("Date: January 25, 2025", body))
    elements.append(PageBreak())

    # Technology Stack
    elements.append(Paragraph("1. Technology Infrastructure Assessment", h1))
    elements.append(Paragraph(
        "The Company's platform is built on a modern cloud-native architecture deployed on AWS. "
        "Key infrastructure observations:", body))

    elements.append(make_table(
        ["Component", "Technology", "Assessment", "Risk Level"],
        [
            ["Cloud Provider", "AWS (us-east-1, us-west-2)", "Multi-AZ; no multi-cloud", "Medium"],
            ["Application", "Python/FastAPI + React", "Modern; well-documented", "Low"],
            ["Database", "PostgreSQL (RDS)", "Managed; automated backups", "Low"],
            ["Search", "Elasticsearch", "Self-managed cluster", "Medium"],
            ["Caching", "Redis (ElastiCache)", "Managed; standard config", "Low"],
            ["CDN", "CloudFront", "Global distribution", "Low"],
            ["CI/CD", "GitHub Actions", "Automated; 95% test coverage", "Low"],
            ["Monitoring", "Datadog", "Full observability stack", "Low"],
            ["Data Pipeline", "Apache Airflow", "Self-managed; 3 engineers", "Medium"],
        ],
    ))

    elements.append(Paragraph(
        "Technical Debt Assessment: Overall technical debt is estimated at 6-8 months of engineering "
        "effort. Primary areas: (1) Elasticsearch cluster requires migration to OpenSearch or managed "
        "service ($180K est.), (2) Legacy API endpoints (v1) still serve 12% of traffic and need "
        "deprecation, (3) Frontend codebase has 14% of components using deprecated React patterns.", body))

    elements.append(PageBreak())

    # Org Structure
    elements.append(Paragraph("2. Organizational Structure", h1))
    elements.append(make_table(
        ["Department", "Headcount", "% of Total", "Avg Tenure", "Open Roles", "Attrition (TTM)"],
        [
            ["Engineering", "98", "31%", "2.8 yrs", "12", "15%"],
            ["Sales", "68", "21%", "1.9 yrs", "8", "22%"],
            ["Customer Success", "52", "16%", "2.4 yrs", "5", "12%"],
            ["Product", "28", "9%", "2.6 yrs", "3", "8%"],
            ["Marketing", "22", "7%", "2.1 yrs", "2", "18%"],
            ["G&A (HR/Finance/Legal)", "32", "10%", "3.1 yrs", "2", "6%"],
            ["Operations/IT", "20", "6%", "2.5 yrs", "1", "10%"],
            ["Total", "320", "100%", "2.5 yrs", "33", "14%"],
        ],
    ))

    elements.append(Paragraph(
        "Key Finding: Sales attrition of 22% is significantly above industry benchmark of 15% for "
        "SaaS companies. Exit interviews indicate compensation (40%), career advancement (30%), and "
        "territory restructuring (20%) as primary drivers. VP Sales has implemented a new comp plan "
        "effective Q1 2025 with higher base/variable split (60/40 vs. previous 50/50).", body))

    # Vendor Dependencies
    elements.append(Paragraph("3. Vendor & Supplier Dependencies", h1))
    elements.append(Paragraph(
        "Critical vendor relationships and concentration risk:", body))

    elements.append(make_table(
        ["Vendor", "Service", "Annual Spend", "Contract Exp", "Switching Cost", "Risk"],
        [
            ["Amazon Web Services", "Infrastructure", "$3.4M", "Dec 2026", "Very High", "High"],
            ["Salesforce", "CRM", "$320K", "Mar 2025", "Medium", "Medium"],
            ["Stripe", "Payments", "$210K", "Month-to-month", "Low", "Low"],
            ["Twilio/SendGrid", "Communications", "$145K", "Annual", "Low", "Low"],
            ["Datadog", "Monitoring", "$180K", "Annual", "Medium", "Low"],
            ["ZoomInfo", "Sales Intelligence", "$95K", "Annual", "Low", "Low"],
        ],
    ))

    elements.append(Paragraph(
        "AWS Concentration Risk: 82% of hosting costs are with AWS, with no multi-cloud capability. "
        "A service disruption in us-east-1 (primary region) would impact all customers until failover "
        "to us-west-2 completes (estimated RTO: 4 hours). Management has budgeted $400K for a "
        "multi-cloud proof-of-concept with GCP in FY2025.", body))

    elements.append(PageBreak())

    # KPIs
    elements.append(Paragraph("4. Operational KPIs & SLAs", h1))
    elements.append(make_table(
        ["KPI", "SLA Target", "FY2023 Actual", "FY2024 Actual", "Trend"],
        [
            ["Platform Uptime", "99.95%", "99.93%", "99.97%", "Improving"],
            ["API Response Time (p95)", "<200ms", "185ms", "162ms", "Improving"],
            ["Support Ticket Resolution", "<4 hours (P1)", "3.2 hrs", "2.8 hrs", "Improving"],
            ["Customer Onboarding Time", "<90 days", "95 days", "82 days", "Improving"],
            ["Feature Release Cadence", "Bi-weekly", "22 releases", "26 releases", "Stable"],
            ["Bug Escape Rate", "<2%", "2.4%", "1.8%", "Improving"],
            ["NPS Score", ">50", "68", "72", "Improving"],
            ["CSAT (Support)", ">90%", "88%", "92%", "Improving"],
        ],
    ))

    # Scalability
    elements.append(Paragraph("5. Scalability Assessment", h1))
    elements.append(Paragraph(
        "The platform's current architecture can support approximately 2x current load (900 customers, "
        "$84M ARR) before requiring significant infrastructure re-architecture. Key bottlenecks:", body))
    elements.append(Paragraph("- Database: Current RDS instance (db.r5.4xlarge) at 45% CPU utilization during peak. "
                              "Horizontal read replicas can extend to ~3x capacity.", bullet))
    elements.append(Paragraph("- Search: Elasticsearch cluster showing strain at current query volumes. Migration "
                              "to OpenSearch Serverless recommended by FY2026.", bullet))
    elements.append(Paragraph("- Application: Stateless microservices auto-scale effectively. No concerns.", bullet))
    elements.append(Paragraph("- Data Pipeline: Airflow DAGs process 2.3TB/day. Current limit is ~4TB/day "
                              "before requiring cluster expansion ($60K/yr incremental).", bullet))
    elements.append(Paragraph("- Estimated infrastructure investment needed for $100M ARR: $1.8M one-time + "
                              "$500K/yr incremental operating costs.", bullet))

    doc.build(elements)
    print(f"  Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("Generating Extended Acme Cloud Solutions DD Documents")
    print("=" * 60)

    generate_qoe_report()
    generate_lbo_model()
    generate_sim()
    generate_legal_summary()
    generate_detailed_financials()
    generate_operational_review()

    print("\n" + "=" * 60)
    print("Done! Generated 6 additional documents")
    print("=" * 60)
