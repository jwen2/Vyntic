"""
Generate realistic sample deal documents for DealInsights PoC testing.
Creates PDFs and Excel files for 3 fictional PE deal targets,
inspired by publicly available financial profiles of mid-market companies.
"""
import os
from pathlib import Path

# ---------------------------------------------------------------------------
# We use reportlab for PDFs and openpyxl for Excel
# ---------------------------------------------------------------------------
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
except ImportError:
    print("Installing reportlab...")
    os.system("pip install reportlab")
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(__file__).parent

# ============================================================================
# DEAL DATA (3 fictional companies with realistic mid-market financials)
# ============================================================================

DEALS = {
    "acme_saas": {
        "name": "Acme Cloud Solutions, Inc.",
        "industry": "B2B SaaS — Enterprise Resource Planning",
        "hq": "Austin, TX",
        "founded": 2014,
        "employees": 320,
        "cim": {
            "executive_summary": (
                "Acme Cloud Solutions is a high-growth B2B SaaS company providing cloud-based "
                "ERP solutions to mid-market manufacturing and distribution companies. The Company "
                "has achieved $42M ARR with 118% net revenue retention, serving over 450 enterprise "
                "customers across North America. Acme's platform integrates supply chain management, "
                "financial planning, and workforce optimization into a single unified interface."
            ),
            "investment_highlights": [
                "Strong unit economics with LTV/CAC ratio of 5.2x",
                "118% net dollar retention driven by platform expansion",
                "85% gross margins with improving operating leverage",
                "Large addressable market ($18B TAM for mid-market ERP)",
                "Experienced management team with prior exits",
                "Minimal customer concentration — top 10 customers represent 22% of ARR",
            ],
            "financial_summary": {
                "headers": ["Metric", "FY2022", "FY2023", "FY2024", "FY2025E"],
                "rows": [
                    ["Revenue ($M)", "24.1", "31.5", "42.0", "54.6"],
                    ["YoY Growth (%)", "38%", "31%", "33%", "30%"],
                    ["Gross Profit ($M)", "19.7", "26.2", "35.7", "46.9"],
                    ["Gross Margin (%)", "82%", "83%", "85%", "86%"],
                    ["EBITDA ($M)", "-2.1", "1.8", "5.9", "10.9"],
                    ["EBITDA Margin (%)", "-9%", "6%", "14%", "20%"],
                    ["Free Cash Flow ($M)", "-4.2", "-0.5", "3.2", "7.8"],
                    ["ARR ($M)", "26.3", "34.0", "42.0", "54.6"],
                    ["Net Revenue Retention", "112%", "115%", "118%", "120%"],
                    ["Customers", "310", "370", "450", "540"],
                    ["CAC Payback (months)", "18", "16", "14", "13"],
                ],
            },
            "key_risks": [
                "Competition from Tier-1 ERP vendors (SAP, Oracle) moving downmarket",
                "Dependency on AWS infrastructure (82% of hosting costs)",
                "Sales cycle averaging 4-6 months for enterprise deals",
                "Key person risk: CTO holds critical architectural knowledge",
            ],
            "change_of_control": (
                "Section 8.4 — Change of Control Provision: In the event of a Change of Control "
                "(defined as the acquisition of more than 50% of voting securities or substantially "
                "all assets), each customer shall have the right to terminate their subscription "
                "agreement within 90 days of written notice. Approximately 35% of ARR ($14.7M) is "
                "subject to contracts containing change of control termination provisions. Management "
                "estimates actual termination risk at less than 5% based on historical precedent."
            ),
            "customer_metrics": {
                "headers": ["Segment", "Customers", "ARR ($M)", "% of Total", "Avg Contract ($K)", "NRR"],
                "rows": [
                    ["Enterprise (>$200K)", "42", "16.8", "40%", "400", "125%"],
                    ["Mid-Market ($50-200K)", "158", "15.8", "38%", "100", "118%"],
                    ["SMB (<$50K)", "250", "9.4", "22%", "38", "108%"],
                ],
            },
        },
        "financials_excel": {
            "income_statement": {
                "headers": ["", "FY2022", "FY2023", "FY2024", "FY2025E"],
                "rows": [
                    ["Subscription Revenue", 21700, 28800, 38600, 50700],
                    ["Professional Services", 2400, 2700, 3400, 3900],
                    ["Total Revenue", 24100, 31500, 42000, 54600],
                    ["", "", "", "", ""],
                    ["Cost of Revenue", 4338, 5355, 6300, 7644],
                    ["Gross Profit", 19762, 26145, 35700, 46956],
                    ["Gross Margin %", "82.0%", "83.0%", "85.0%", "86.0%"],
                    ["", "", "", "", ""],
                    ["Sales & Marketing", 10845, 12600, 14700, 17472],
                    ["Research & Development", 7230, 8190, 10500, 12012],
                    ["General & Administrative", 3856, 3528, 4620, 6006],
                    ["Total OpEx", 21931, 24318, 29820, 35490],
                    ["", "", "", "", ""],
                    ["EBITDA", -2169, 1827, 5880, 11466],
                    ["EBITDA Margin %", "-9.0%", "5.8%", "14.0%", "21.0%"],
                    ["D&A", 482, 630, 840, 1092],
                    ["EBIT", -2651, 1197, 5040, 10374],
                    ["Interest Expense", 120, 95, 80, 65],
                    ["Pre-Tax Income", -2771, 1102, 4960, 10309],
                    ["Tax Provision", 0, 165, 1240, 2577],
                    ["Net Income", -2771, 937, 3720, 7732],
                ],
            },
            "balance_sheet": {
                "headers": ["", "FY2022", "FY2023", "FY2024"],
                "rows": [
                    ["Cash & Equivalents", 8200, 12500, 18700],
                    ["Accounts Receivable", 4100, 5200, 6800],
                    ["Prepaid Expenses", 850, 1100, 1400],
                    ["Total Current Assets", 13150, 18800, 26900],
                    ["", "", "", ""],
                    ["PP&E (net)", 1200, 1500, 1800],
                    ["Intangible Assets", 3200, 2800, 2400],
                    ["Goodwill", 5600, 5600, 5600],
                    ["Total Assets", 23150, 28700, 36700],
                    ["", "", "", ""],
                    ["Accounts Payable", 1800, 2200, 2800],
                    ["Deferred Revenue", 6400, 8500, 10500],
                    ["Accrued Expenses", 2100, 2800, 3500],
                    ["Total Current Liabilities", 10300, 13500, 16800],
                    ["", "", "", ""],
                    ["Long-term Debt", 2000, 1500, 1000],
                    ["Total Liabilities", 12300, 15000, 17800],
                    ["", "", "", ""],
                    ["Total Equity", 10850, 13700, 18900],
                ],
            },
        },
    },

    "pinnacle_health": {
        "name": "Pinnacle Healthcare Services, LLC",
        "industry": "Healthcare Services — Outpatient Rehabilitation",
        "hq": "Nashville, TN",
        "founded": 2009,
        "employees": 1850,
        "cim": {
            "executive_summary": (
                "Pinnacle Healthcare Services is a leading provider of outpatient physical therapy "
                "and rehabilitation services operating 78 clinics across the Southeastern United States. "
                "The Company generated $127M in revenue and $19.1M in Adjusted EBITDA for the trailing "
                "twelve months ending Q3 2024. Pinnacle has completed 12 tuck-in acquisitions since 2019, "
                "building a dense regional network with strong referral relationships with over 2,200 "
                "referring physicians."
            ),
            "investment_highlights": [
                "Defensive healthcare services with non-discretionary demand",
                "78 clinics with proven de novo and M&A growth playbook",
                "15% Adjusted EBITDA margin with pathway to 18%+",
                "Diversified payor mix: 45% commercial, 32% Medicare, 23% other",
                "Strong same-store growth of 6.2% driven by volume and rate increases",
                "Fragmented market with 50,000+ independent PT clinics nationally",
            ],
            "financial_summary": {
                "headers": ["Metric", "FY2022", "FY2023", "FY2024E", "FY2025E"],
                "rows": [
                    ["Revenue ($M)", "98.2", "112.5", "127.0", "146.1"],
                    ["YoY Growth (%)", "12%", "15%", "13%", "15%"],
                    ["Gross Profit ($M)", "53.0", "61.9", "71.1", "83.3"],
                    ["Gross Margin (%)", "54%", "55%", "56%", "57%"],
                    ["Adjusted EBITDA ($M)", "12.7", "15.8", "19.1", "24.8"],
                    ["EBITDA Margin (%)", "13%", "14%", "15%", "17%"],
                    ["Capex ($M)", "5.9", "7.9", "8.9", "10.2"],
                    ["Free Cash Flow ($M)", "4.8", "6.2", "8.5", "12.4"],
                    ["Clinics", "58", "68", "78", "92"],
                    ["Visits (000s)", "890", "1020", "1150", "1340"],
                    ["Revenue per Visit ($)", "110", "110", "110", "109"],
                    ["Patient Satisfaction (NPS)", "72", "74", "76", "78"],
                ],
            },
            "key_risks": [
                "Medicare reimbursement rate changes (CMS proposed -3.4% cut for 2025)",
                "Therapist recruitment and retention in tight labor market",
                "Integration risk from rapid acquisition pace",
                "Geographic concentration in Southeast (82% of revenue)",
            ],
            "change_of_control": (
                "Section 12.1 — Assignment and Change of Control: This Agreement may not be assigned "
                "by either party without prior written consent. In the event of a Change of Control "
                "(defined as a merger, consolidation, or sale of all or substantially all assets), "
                "payor contracts representing approximately $41M in annual revenue (32% of total) "
                "require payor notification and consent within 60 days. Medicare provider enrollment "
                "must be updated with CMS within 90 days of any ownership change exceeding 5%."
            ),
            "clinic_performance": {
                "headers": ["Region", "Clinics", "Revenue ($M)", "EBITDA ($M)", "Margin", "Avg Visits/Day"],
                "rows": [
                    ["Tennessee", "22", "38.1", "6.5", "17%", "32"],
                    ["Georgia", "18", "27.9", "4.5", "16%", "29"],
                    ["Alabama", "14", "20.3", "3.0", "15%", "27"],
                    ["Florida", "12", "22.9", "3.2", "14%", "35"],
                    ["Carolinas", "12", "17.8", "1.9", "11%", "24"],
                ],
            },
        },
        "financials_excel": {
            "income_statement": {
                "headers": ["", "FY2022", "FY2023", "FY2024E", "FY2025E"],
                "rows": [
                    ["Patient Service Revenue", 89200, 102400, 115600, 133400],
                    ["Management Fees", 5800, 6500, 7200, 8100],
                    ["Other Revenue", 3200, 3600, 4200, 4600],
                    ["Total Revenue", 98200, 112500, 127000, 146100],
                    ["", "", "", "", ""],
                    ["Direct Costs (Labor)", 34370, 38250, 42418, 48213],
                    ["Supplies & Occupancy", 10802, 12375, 13462, 14610],
                    ["Total Cost of Services", 45172, 50625, 55868, 62823],
                    ["Gross Profit", 53028, 61875, 71132, 83277],
                    ["Gross Margin %", "54.0%", "55.0%", "56.0%", "57.0%"],
                    ["", "", "", "", ""],
                    ["Clinic Operating Expenses", 24550, 25875, 27940, 29220],
                    ["Corporate G&A", 9820, 10125, 11430, 11688],
                    ["Acquisition Costs", 1964, 3375, 2540, 4383],
                    ["Total OpEx", 36334, 39375, 41910, 45291],
                    ["", "", "", "", ""],
                    ["Adjusted EBITDA", 12694, 15750, 19050, 24786],
                    ["EBITDA Margin %", "12.9%", "14.0%", "15.0%", "17.0%"],
                    ["D&A", 4910, 5625, 6350, 7305],
                    ["EBIT", 7784, 10125, 12700, 17481],
                    ["Interest Expense", 2946, 3375, 3810, 4383],
                    ["Pre-Tax Income", 4838, 6750, 8890, 13098],
                    ["Tax Provision", 1209, 1688, 2222, 3275],
                    ["Net Income", 3629, 5062, 6668, 9823],
                ],
            },
            "balance_sheet": {
                "headers": ["", "FY2022", "FY2023", "FY2024E"],
                "rows": [
                    ["Cash & Equivalents", 4200, 5800, 8200],
                    ["Accounts Receivable", 14700, 16900, 19050],
                    ["Prepaid & Other", 2800, 3200, 3600],
                    ["Total Current Assets", 21700, 25900, 30850],
                    ["", "", "", ""],
                    ["PP&E (net)", 18200, 22400, 26800],
                    ["Intangible Assets", 8500, 11200, 13600],
                    ["Goodwill", 42000, 54000, 62000],
                    ["Other Long-term Assets", 3200, 3800, 4200],
                    ["Total Assets", 93600, 117300, 137450],
                    ["", "", "", ""],
                    ["Accounts Payable", 5900, 6800, 7600],
                    ["Accrued Compensation", 4200, 4800, 5400],
                    ["Current Portion of Debt", 3200, 3600, 4000],
                    ["Total Current Liabilities", 13300, 15200, 17000],
                    ["", "", "", ""],
                    ["Long-term Debt", 38000, 48000, 52000],
                    ["Other Long-term Liabilities", 4800, 5600, 6200],
                    ["Total Liabilities", 56100, 68800, 75200],
                    ["", "", "", ""],
                    ["Total Equity", 37500, 48500, 62250],
                ],
            },
        },
    },

    "summit_industrial": {
        "name": "Summit Precision Manufacturing, Inc.",
        "industry": "Specialty Industrial — Aerospace & Defense Components",
        "hq": "Wichita, KS",
        "founded": 1987,
        "employees": 680,
        "cim": {
            "executive_summary": (
                "Summit Precision Manufacturing is a leading producer of mission-critical machined "
                "components and sub-assemblies for the aerospace and defense industry. The Company "
                "operates three ISO 9001/AS9100-certified facilities totaling 185,000 sq ft and "
                "serves Tier-1 OEMs including Boeing, Lockheed Martin, and Raytheon. Summit generated "
                "$89M in revenue with $16.9M in EBITDA (19% margin) in FY2024, supported by a $142M "
                "backlog providing 18+ months of revenue visibility."
            ),
            "investment_highlights": [
                "Mission-critical components with high switching costs and long qualification cycles",
                "19% EBITDA margin with expansion opportunity through automation investment",
                "$142M backlog providing exceptional revenue visibility",
                "Dual-use platform serving both commercial aerospace and defense",
                "ITAR-registered and NADCAP-accredited — significant barriers to entry",
                "Secular tailwind: global defense spending and commercial aircraft production ramp",
            ],
            "financial_summary": {
                "headers": ["Metric", "FY2022", "FY2023", "FY2024", "FY2025E"],
                "rows": [
                    ["Revenue ($M)", "72.4", "79.8", "89.0", "98.0"],
                    ["YoY Growth (%)", "8%", "10%", "12%", "10%"],
                    ["Gross Profit ($M)", "26.8", "30.3", "34.7", "39.2"],
                    ["Gross Margin (%)", "37%", "38%", "39%", "40%"],
                    ["EBITDA ($M)", "11.6", "13.6", "16.9", "19.6"],
                    ["EBITDA Margin (%)", "16%", "17%", "19%", "20%"],
                    ["Capex ($M)", "4.3", "5.6", "7.1", "6.9"],
                    ["Free Cash Flow ($M)", "5.8", "6.4", "8.2", "11.0"],
                    ["Backlog ($M)", "95.0", "118.0", "142.0", "155.0"],
                    ["Book-to-Bill Ratio", "1.08", "1.12", "1.15", "1.10"],
                    ["On-Time Delivery (%)", "94%", "95%", "96%", "97%"],
                    ["Churn Rate (%)", "3.2%", "2.8%", "2.1%", "1.8%"],
                ],
            },
            "key_risks": [
                "Customer concentration: Boeing and Lockheed represent 48% of revenue",
                "Titanium and specialty alloy price volatility",
                "Skilled machinist labor shortage (avg age of workforce: 52 years)",
                "Single-source risk on 3 CNC machine tool suppliers",
            ],
            "change_of_control": (
                "Section 15.3 — Change of Control: Under the Master Supply Agreements with Boeing "
                "(representing $28M annual revenue) and Lockheed Martin ($14.5M annual revenue), "
                "a Change of Control requires written notification within 30 days and allows the "
                "customer to audit the acquirer's financial condition and manufacturing capabilities. "
                "Either customer may terminate with 180 days notice if the acquirer fails to maintain "
                "AS9100 certification or ITAR registration. No accelerated payment provisions exist. "
                "All government contracts (18% of revenue) are subject to novation approval by the "
                "relevant Contracting Officer, typically requiring 6-12 months."
            ),
            "revenue_by_program": {
                "headers": ["Program/Customer", "Revenue ($M)", "% of Total", "Contract Type", "Expiry"],
                "rows": [
                    ["Boeing 737 MAX", "18.5", "21%", "LTA", "2027"],
                    ["Boeing 787", "9.6", "11%", "LTA", "2026"],
                    ["Lockheed F-35", "14.5", "16%", "IDIQ", "2029"],
                    ["Raytheon Missiles", "8.9", "10%", "Fixed Price", "2026"],
                    ["GE Aerospace", "7.1", "8%", "LTA", "2027"],
                    ["Other Commercial", "18.4", "21%", "Various", "Various"],
                    ["Other Defense", "12.0", "13%", "Various", "Various"],
                ],
            },
        },
        "financials_excel": {
            "income_statement": {
                "headers": ["", "FY2022", "FY2023", "FY2024", "FY2025E"],
                "rows": [
                    ["Commercial Aerospace", 39100, 43100, 49500, 54900],
                    ["Defense", 22470, 24740, 26700, 29400],
                    ["Industrial / Other", 10830, 11960, 12800, 13700],
                    ["Total Revenue", 72400, 79800, 89000, 98000],
                    ["", "", "", "", ""],
                    ["Direct Materials", 25340, 27132, 29370, 31360],
                    ["Direct Labor", 13756, 14364, 16020, 17640],
                    ["Manufacturing Overhead", 6516, 7980, 8900, 9800],
                    ["Total COGS", 45612, 49476, 54290, 58800],
                    ["Gross Profit", 26788, 30324, 34710, 39200],
                    ["Gross Margin %", "37.0%", "38.0%", "39.0%", "40.0%"],
                    ["", "", "", "", ""],
                    ["Sales & Marketing", 3620, 3990, 4005, 4410],
                    ["Engineering", 5068, 5586, 6230, 6860],
                    ["G&A", 6516, 7182, 7565, 8330],
                    ["Total OpEx", 15204, 16758, 17800, 19600],
                    ["", "", "", "", ""],
                    ["EBITDA", 11584, 13566, 16910, 19600],
                    ["EBITDA Margin %", "16.0%", "17.0%", "19.0%", "20.0%"],
                    ["D&A", 4344, 4788, 5340, 5880],
                    ["EBIT", 7240, 8778, 11570, 13720],
                    ["Interest Expense", 1448, 1596, 1780, 1470],
                    ["Pre-Tax Income", 5792, 7182, 9790, 12250],
                    ["Tax Provision", 1448, 1796, 2448, 3063],
                    ["Net Income", 4344, 5386, 7342, 9187],
                ],
            },
            "balance_sheet": {
                "headers": ["", "FY2022", "FY2023", "FY2024"],
                "rows": [
                    ["Cash & Equivalents", 3800, 4200, 6500],
                    ["Accounts Receivable", 12100, 13300, 14800],
                    ["Inventory", 8900, 10200, 11400],
                    ["Total Current Assets", 24800, 27700, 32700],
                    ["", "", "", ""],
                    ["PP&E (net)", 28500, 32200, 36800],
                    ["Intangible Assets", 4200, 3800, 3400],
                    ["Goodwill", 12000, 12000, 12000],
                    ["Total Assets", 69500, 75700, 84900],
                    ["", "", "", ""],
                    ["Accounts Payable", 6500, 7200, 8000],
                    ["Accrued Liabilities", 4800, 5300, 5900],
                    ["Current Portion of Debt", 2400, 2400, 2400],
                    ["Total Current Liabilities", 13700, 14900, 16300],
                    ["", "", "", ""],
                    ["Long-term Debt", 18000, 16200, 14400],
                    ["Other Long-term Liabilities", 3200, 3400, 3600],
                    ["Total Liabilities", 34900, 34500, 34300],
                    ["", "", "", ""],
                    ["Total Equity", 34600, 41200, 50600],
                ],
            },
        },
    },
}


# ============================================================================
# PDF GENERATION
# ============================================================================

def make_table_flowable(headers, rows, col_widths=None):
    """Create a reportlab Table with styling."""
    data = [headers] + rows
    if col_widths is None:
        col_widths = [1.8 * inch] + [1.1 * inch] * (len(headers) - 1)
        # Adjust to fit page
        total = sum(col_widths)
        if total > 6.5 * inch:
            ratio = 6.5 * inch / total
            col_widths = [w * ratio for w in col_widths]

    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return t


def generate_cim_pdf(deal_id: str, deal: dict):
    """Generate a CIM PDF for a deal."""
    filename = OUT / f"{deal_id}_cim.pdf"
    doc = SimpleDocTemplate(str(filename), pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("DealTitle", parent=styles["Title"],
                                  fontSize=18, textColor=colors.HexColor("#1a365d"))
    h1 = ParagraphStyle("H1", parent=styles["Heading1"],
                         fontSize=14, textColor=colors.HexColor("#2d3748"),
                         spaceBefore=18, spaceAfter=8)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"],
                         fontSize=11, textColor=colors.HexColor("#4a5568"),
                         spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle("Body", parent=styles["Normal"],
                           fontSize=9.5, leading=14, spaceBefore=4, spaceAfter=4)
    bullet = ParagraphStyle("Bullet", parent=body, leftIndent=20,
                             bulletIndent=10, spaceBefore=2, spaceAfter=2)
    conf = ParagraphStyle("Conf", parent=styles["Normal"], fontSize=8,
                           textColor=colors.red, alignment=1)

    cim = deal["cim"]
    elements = []

    # Title page
    elements.append(Spacer(1, 2 * inch))
    elements.append(Paragraph("CONFIDENTIAL INFORMATION MEMORANDUM", conf))
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph(deal["name"], title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"Industry: {deal['industry']}", body))
    elements.append(Paragraph(f"Headquarters: {deal['hq']}", body))
    elements.append(Paragraph(f"Founded: {deal['founded']}  |  Employees: {deal['employees']}", body))
    elements.append(Spacer(1, 0.5 * inch))
    elements.append(Paragraph("Prepared for Qualified Institutional Buyers", body))
    elements.append(PageBreak())

    # Executive Summary
    elements.append(Paragraph("1. Executive Summary", h1))
    elements.append(Paragraph(cim["executive_summary"], body))

    # Investment Highlights
    elements.append(Paragraph("2. Investment Highlights", h1))
    for item in cim["investment_highlights"]:
        elements.append(Paragraph(f"• {item}", bullet))

    # Financial Summary
    elements.append(Paragraph("3. Financial Summary", h1))
    fs = cim["financial_summary"]
    elements.append(make_table_flowable(fs["headers"], fs["rows"]))

    elements.append(PageBreak())

    # Additional metrics table
    if "customer_metrics" in cim:
        elements.append(Paragraph("4. Customer Segmentation", h1))
        cm = cim["customer_metrics"]
        elements.append(make_table_flowable(cm["headers"], cm["rows"]))
    elif "clinic_performance" in cim:
        elements.append(Paragraph("4. Clinic Performance by Region", h1))
        cp = cim["clinic_performance"]
        elements.append(make_table_flowable(cp["headers"], cp["rows"]))
    elif "revenue_by_program" in cim:
        elements.append(Paragraph("4. Revenue by Program", h1))
        rp = cim["revenue_by_program"]
        elements.append(make_table_flowable(rp["headers"], rp["rows"]))

    # Key Risks
    elements.append(Paragraph("5. Key Risks", h1))
    for item in cim["key_risks"]:
        elements.append(Paragraph(f"• {item}", bullet))

    # Change of Control
    elements.append(Paragraph("6. Change of Control Provisions", h1))
    elements.append(Paragraph(cim["change_of_control"], body))

    doc.build(elements)
    print(f"  ✓ Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def generate_financials_xlsx(deal_id: str, deal: dict):
    """Generate a financial model Excel file for a deal."""
    filename = OUT / f"{deal_id}_financials.xlsx"
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    label_font = Font(name="Calibri", bold=True, size=10)
    number_font = Font(name="Calibri", size=10)
    pct_font = Font(name="Calibri", size=10, italic=True, color="4A5568")
    thin_border = Border(
        bottom=Side(style="thin", color="CBD5E0"),
    )

    for sheet_name, sheet_data in deal["financials_excel"].items():
        ws = wb.create_sheet(title=sheet_name.replace("_", " ").title())
        headers = sheet_data["headers"]
        rows = sheet_data["rows"]

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1,
                             value=f"{deal['name']} — {sheet_name.replace('_', ' ').title()}")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1A365D")

        # Headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

        # Data
        for row_idx, row in enumerate(rows, 4):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                if col_idx == 1:
                    cell.font = label_font if val and not str(val).endswith("%") else pct_font
                elif isinstance(val, str) and "%" in val:
                    cell.font = pct_font
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(val, (int, float)) and val != "":
                    cell.font = number_font
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

        # Column widths
        ws.column_dimensions["A"].width = 28
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 14

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(filename))
    print(f"  ✓ Created {filename.name} ({filename.stat().st_size // 1024}KB)")
    return filename


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("Generating DealInsights Sample Deal Documents")
    print("=" * 60)

    for deal_id, deal in DEALS.items():
        print(f"\n📁 {deal['name']} (deal_id: {deal_id})")
        generate_cim_pdf(deal_id, deal)
        generate_financials_xlsx(deal_id, deal)

    print("\n" + "=" * 60)
    print(f"Done! Generated {len(DEALS) * 2} files in {OUT}")
    print("=" * 60)
